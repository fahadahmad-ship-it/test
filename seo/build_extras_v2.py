#!/usr/bin/env python3
"""Build premium standalone Disavow + Baseline workbooks (match audit design)."""
import os, collections, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("bv", os.path.join(BASE, "build_audit_v2.py"))
B = importlib.util.module_from_spec(spec); spec.loader.exec_module(B)

def build():
    refs, bls = B.get_data()
    toxic = [r for r in refs if r["verdict"] == "TOXIC"]
    keep = [r for r in refs if r["verdict"] == "KEEP"]
    monitor = [r for r in refs if r["verdict"] == "MONITOR"]
    own = [r for r in refs if r["verdict"] == "OWN"]
    total = len(refs)
    pct = f"{len(toxic)/total*100:.1f}%"

    # ================= DISAVOW WORKBOOK =================
    wb = Workbook()
    ws = wb.active; ws.title = "Disavow List"
    ws.sheet_view.showGridLines = False
    cols = ["#", "Disavow entry", "Referring domain", "Authority Score", "Country", "Reason / evidence"]
    widths = [5, 34, 30, 10, 8, 62]
    # KPI tile band above the table
    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = f"DISAVOW LIST — {len(toxic)} domains (QA-PENDING)"
    c.font = B.F(15, True, B.WHITE); c.fill = B.fill(B.NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    s = ws["A2"]; s.value = ("Domain-level entries for Google's Disavow tool (copy column B into the .txt). "
                             "Excludes KEEP / MONITOR / OWN. Do NOT submit until false-positive sweep + client sign-off.")
    s.font = B.F(10, False, B.INK_SOFT, italic=True); s.alignment = Alignment(indent=1)
    B.header_row(ws, cols, 3, widths)
    n = 0
    for rr in sorted(toxic, key=lambda x: (-x["tox"], x["domain"])):
        n += 1
        ws.append([n, f"domain:{rr['domain']}", rr["domain"], rr["ascore"], rr["country"], rr["reason"]])
        if n % 2 == 0:
            for cc in range(1, len(cols) + 1):
                if not ws.cell(row=ws.max_row, column=cc).fill.patternType:
                    ws.cell(row=ws.max_row, column=cc).fill = B.fill(B.VC_SOFT["TOXIC"])
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{ws.max_row}"

    ws2 = wb.create_sheet("README")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["B"].width = 100
    B.banner(ws2, "DISAVOW — HOW TO USE", 2)
    lines = [
        ("STATUS: QA-PENDING — do not upload until sign-off.", True),
        ("", False),
        (f"Included: {len(toxic)} referring domains classified TOXIC across two QA passes.", False),
        ("  PBN / directory / article / bookmark link-farms (single-IP clusters)", False),
        ("  SEO / link-selling / 'buy backlinks' / rank schemes", False),
        ("  gambling / adult / off-topic high-risk verticals", False),
        ("  foreign-hosting spam clusters with English lawn-care anchors", False),
        ("", False),
        ("Excluded on purpose:", True),
        ("  OWN client/sister properties (masterlawn.org/.net, masterlawninc.com, midsouthturf, greenkingspray)", False),
        (f"  KEEP ({len(keep)}) legitimate/relevant + MONITOR ({len(monitor)}) watch-only", False),
        ("", False),
        ("Submission steps:", True),
        ("  1. Copy column B (domain: lines) into disavow_masterlawn.txt (UTF-8).", False),
        ("  2. Validate format; false-positive sweep vs KEEP/MONITOR + GSC 'Top linking sites'.", False),
        ("  3. Two-analyst + client sign-off.", False),
        ("  4. Upload to the correct GSC property via the Disavow Links tool.", False),
        ("  5. Refresh MONTHLY (injection ongoing) and always re-upload the full cumulative file.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        cc = ws2.cell(row=i + 2, column=2, value=t)
        cc.font = B.F(10, b, B.NAVY if b else B.INK)
    wb.save(os.path.join(BASE, "MasterLawn_Disavow.xlsx"))

    # ================= BASELINE WORKBOOK =================
    wb = Workbook()
    ws = wb.active; ws.title = "Baseline Benchmark"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([3, 20, 16, 16, 16, 16, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells("B2:I2")
    t = ws["B2"]; t.value = "MASTER LAWN — CRITICAL SEO BASELINE BENCHMARK"
    t.font = B.F(18, True, B.WHITE); t.fill = B.fill(B.NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 34
    ws.merge_cells("B3:I3")
    st = ws["B3"]; st.value = "Source: Semrush (Authority Score & Traffic)  ·  Snapshot 2026-09-04  ·  Frozen 'before' state for measuring disavow + rebuild impact"
    st.font = B.F(10, False, B.INK_SOFT, italic=True)

    B.tile(ws, 5, 2, "AUTHORITY SCORE", 23)
    B.tile(ws, 5, 4, "REFERRING DOMAINS", 1167)
    B.tile(ws, 5, 6, "TOTAL BACKLINKS", 3613)
    B.tile(ws, 5, 8, "% TOXIC", pct, valcolor=B.RED, fillc=B.VC_SOFT["TOXIC"])
    B.tile(ws, 9, 2, "ORGANIC KEYWORDS", 2525)
    B.tile(ws, 9, 4, "ORGANIC TRAFFIC / MO", 2391)
    B.tile(ws, 9, 6, "TRAFFIC VALUE / MO", "$10,324")
    B.tile(ws, 9, 8, "TOXIC DOMAINS", len(toxic), valcolor=B.RED, fillc=B.VC_SOFT["TOXIC"])

    r = 13
    def section(title):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        c = ws.cell(row=r, column=2, value=title); c.font = B.F(12, True, B.NAVY); c.fill = B.fill(B.BAND)
        r += 1
    def kv(k, v, note=""):
        nonlocal r
        a = ws.cell(row=r, column=2, value=k); a.font = B.F(10)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        b = ws.cell(row=r, column=3, value=v); b.font = B.F(10, True, B.NAVY)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        c = ws.cell(row=r, column=5, value=note); c.font = B.F(9, False, B.INK_SOFT)
        for cc in (2, 3, 5): ws.cell(row=r, column=cc).border = B.BORDER
        r += 1

    section("AUTHORITY & BACKLINK PROFILE (Semrush)")
    for k, v, nt in [("Authority Score", 23, "inflated by spam volume, not earned trust"),
                     ("Total backlinks", 3613, ""), ("Referring domains", 1167, "~90% at Authority Score ≤6"),
                     ("Referring IPs (Class C)", "1,151 (792)", ""), ("Follow / Nofollow", "2,877 / 743", ""),
                     ("Text / Image links", "3,243 / 104", "")]:
        kv(k, v, nt)
    r += 1
    section("TOXICITY (this audit, 2 QA passes)")
    for k, v, nt in [("TOXIC — disavow", len(toxic), pct + " of referring domains"),
                     ("KEEP", len(keep), "legitimate / relevant"), ("MONITOR", len(monitor), "watch only"),
                     ("OWN", own.__len__(), "client / sister properties"),
                     ("Recency", "~94% first seen 2026", "active negative-SEO blast")]:
        kv(k, v, nt)
    r += 1
    section("ORGANIC VISIBILITY (Semrush, US)")
    for k, v, nt in [("Semrush Rank", 595080, ""), ("Organic keywords", 2525, "104 in top 3"),
                     ("Organic traffic / mo", 2391, ""), ("Organic traffic value / mo", "$10,324", ""),
                     ("Paid keywords / traffic", "0 / 0", "")]:
        kv(k, v, nt)
    r += 1
    section("PENDING CLIENT ACCESS (source of truth once granted)")
    for k, v, nt in [("Google Search Console", "clicks / impressions / CTR / position", "request access"),
                     ("GA4", "organic sessions & conversions", "request access"),
                     ("Rank tracker (local pack)", "Semrush Position Tracking", "to configure"),
                     ("Confirm ownership", "masterlawninc / .org / .net / midsouthturf / greenkingspray", "")]:
        kv(k, v, nt)

    ws = wb.create_sheet("Local Keyword Set")
    ws.sheet_view.showGridLines = False
    B.header_row(ws, ["Service", "Geo modifiers to track"], B.banner(ws, "LOCAL KEYWORD SET", 2), [24, 88])
    geos = "Memphis TN; Germantown TN; Collierville TN; Bartlett TN; Olive Branch MS; Southaven MS; Huntsville AL"
    for s in ["lawn care", "lawn fertilization", "weed control", "lawn aeration", "mosquito control",
              "lawn treatment", "lawn maintenance"]:
        ws.append([s, geos])

    ws = wb.create_sheet("Change Log")
    ws.sheet_view.showGridLines = False
    B.header_row(ws, ["Date", "Authority Score", "Ref domains", "Toxic", "Organic KW", "Traffic/mo", "Note"],
                 B.banner(ws, "CHANGE LOG (add a row monthly)", 7), [12, 14, 12, 9, 11, 11, 40])
    ws.append(["2026-09-04", 23, 1167, len(toxic), 2525, 2391, "Baseline captured (pre-disavow)"])
    wb.save(os.path.join(BASE, "MasterLawn_Baseline_Benchmark.xlsx"))

    # refresh plain-text disavow to match
    with open(os.path.join(BASE, "disavow_masterlawn.txt"), "w", encoding="utf-8") as f:
        import datetime
        f.write("# Master Lawn — disavow file (QA-PENDING, do not submit until signed off)\n")
        f.write(f"# Generated {datetime.date.today()} from Semrush backlink audit (2 QA passes).\n")
        f.write(f"# {len(toxic)} toxic referring domains (domain-level).\n#\n")
        for rr in sorted(toxic, key=lambda x: x["domain"]):
            f.write(f"domain:{rr['domain']}\n")

    print(f"Disavow + Baseline rebuilt. toxic={len(toxic)} keep={len(keep)} monitor={len(monitor)} own={len(own)}")

if __name__ == "__main__":
    build()
