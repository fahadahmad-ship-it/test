#!/usr/bin/env python3
"""
Master Lawn — premium backlink-audit workbook builder (v2).
Dashboard + charts, executive summary with KPI tiles, banner-styled numbered
tables with reason column, and an EMBEDDED disavow list.
Reuses the classifier + QA overrides from classify.py.
"""
import os, csv, collections, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties

BASE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("classify", os.path.join(BASE, "classify.py"))
C = importlib.util.module_from_spec(spec); spec.loader.exec_module(C)

# ---------- palette (validated, dataviz skill light mode) ----------
INK        = "0B0B0B"
INK_SOFT   = "52514E"
NAVY       = "0D366B"   # header/banner fill
NAVY2      = "184F95"
ACCENT     = "2A78D6"   # blue
AQUA       = "1BAF7A"
YELLOW     = "EDA100"
RED        = "E34948"
VIOLET     = "4A3AA7"
ORANGE     = "EB6834"
SURFACE    = "FCFCFB"
BAND       = "F4F6F9"
TILE_BLUE  = "E8F0FB"
# verdict status colors (shipped with labels)
VC = {"TOXIC": RED, "KEEP": AQUA, "MONITOR": YELLOW, "OWN": ACCENT, "UNMAPPED": "BFBFBF"}
VC_SOFT = {"TOXIC": "F8D7D5", "KEEP": "D6F0E6", "MONITOR": "FCEFC7", "OWN": "DBE9FB"}

WHITE = "FFFFFF"
thin = Side(style="thin", color="D9DEE6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(sz=11, b=False, color=INK, italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

# ---------------- load + classify (with all QA overrides) ----------------
def get_data():
    refs = C.load_refdomains(); bls = C.load_backlinks()
    overrides = {}
    files = [os.path.join(BASE, "data", "qa", fn) for fn in
             ("monitor_A_reviewed.csv", "monitor_B_reviewed.csv", "keep_reviewed.csv")]
    files += [os.path.join(BASE, "data", "qa2", f"toxic_{i}_reviewed.csv") for i in range(1, 8)]
    files.append(os.path.join(BASE, "data", "qa2", "keep_final_reviewed.csv"))
    files.append(os.path.join(BASE, "data", "qa3", "local_citations_reviewed.csv"))
    for p in files:
        if not os.path.exists(p): continue
        with open(p, encoding="utf-8", errors="replace") as f:
            rd = csv.reader(f); next(rd, None)
            for row in rd:
                if len(row) < 2 or not row[0].strip(): continue
                fv = row[1].strip().upper()
                if fv in ("TOXIC", "MONITOR", "KEEP"):
                    overrides[C.root_domain(row[0])] = (fv, row[3].strip() if len(row) > 3 else "")
    ipc = collections.Counter(C.ip_prefix(r["ip"], 3) for r in refs)
    for r in refs:
        v, score, code, reasons = C.classify(r["domain"], r["ascore"], r["backlinks"], r["ip"], r["country"])
        ov = overrides.get(r["domain"].lower())
        note = ""
        if ov and v != "OWN":
            if ov[0] != v:
                reasons = reasons + [f"QA→{ov[0]}" + (f" ({ov[1]})" if ov[1] else "")]
            v = ov[0]; note = ov[1]
        r["verdict"] = v; r["tox"] = score
        r["reason"] = "; ".join(reasons)
        r["ip_shared"] = ipc[C.ip_prefix(r["ip"], 3)]
    by = {r["domain"]: r for r in refs}
    for b in bls:
        rr = by.get(b["refdomain"])
        b["verdict"] = rr["verdict"] if rr else "UNMAPPED"
    return refs, bls

# ---------------- helpers ----------------
def banner(ws, text, ncols, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = F(15, True, WHITE); c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    hrow = 2
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=1, value=sub)
        s.font = F(10, False, INK_SOFT, italic=True)
        s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        hrow = 3
    return hrow

def header_row(ws, cols, row, widths=None):
    for i, name in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = F(10, True, WHITE); c.fill = fill(NAVY2)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = f"A{row + 1}"   # string coord: don't instantiate a blank data row
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def tile(ws, r, c, label, value, valcolor=NAVY, fillc=TILE_BLUE, span=2):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 2, end_column=c + span - 1)
    lc = ws.cell(row=r, column=c, value=label)
    lc.font = F(9, True, INK_SOFT); lc.fill = fill(fillc)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws.cell(row=r + 1, column=c, value=value)
    vc.font = F(24, True, valcolor); vc.fill = fill(fillc)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    for rr in range(r, r + 3):
        for cc in range(c, c + span):
            ws.cell(row=rr, column=cc).border = BORDER
            if not ws.cell(row=rr, column=cc).fill.patternType:
                ws.cell(row=rr, column=cc).fill = fill(fillc)

def style_barchart(ch, color):
    ch.legend = None
    ch.y_axis.majorGridlines = None
    ch.x_axis.delete = False; ch.y_axis.delete = False
    ch.x_axis.majorTickMark = "out"; ch.y_axis.majorTickMark = "out"
    for s in ch.series:
        s.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color

# ---------------- build ----------------
def build():
    refs, bls = get_data()
    verd = collections.Counter(r["verdict"] for r in refs)
    toxic = [r for r in refs if r["verdict"] == "TOXIC"]
    keep = [r for r in refs if r["verdict"] == "KEEP"]
    monitor = [r for r in refs if r["verdict"] == "MONITOR"]
    own = [r for r in refs if r["verdict"] == "OWN"]
    total = len(refs)

    # distributions
    def bucket_as(a):
        for lo in (0, 10, 20, 30, 40, 50, 70):
            pass
        if a <= 9: return "0–9"
        if a <= 19: return "10–19"
        if a <= 29: return "20–29"
        if a <= 39: return "30–39"
        if a <= 49: return "40–49"
        if a <= 69: return "50–69"
        return "70+"
    asb = collections.Counter(bucket_as(r["ascore"]) for r in refs)
    asb_order = ["0–9","10–19","20–29","30–39","40–49","50–69","70+"]
    country = collections.Counter((r["country"].upper() if r["country"] else "Unknown") for r in refs)
    tld = collections.Counter(r["domain"].rsplit(".",1)[-1] for r in refs)
    def yr(d): return (d or "")[:4]
    years = collections.Counter(yr(r["first_seen"]) for r in toxic if yr(r["first_seen"]))
    ipfarm = collections.Counter(C.ip_prefix(r["ip"],3) for r in toxic)
    all_years = collections.Counter(yr(r["first_seen"]) for r in refs if yr(r["first_seen"]))
    pct2026 = all_years.get("2026", 0) / total * 100

    wb = Workbook()

    # ============ CHART DATA (hidden) ============
    cd = wb.active; cd.title = "Chart Data"
    def put_table(startcol, title, pairs):
        cd.cell(row=1, column=startcol, value=title).font = F(9, True)
        cd.cell(row=2, column=startcol, value="label")
        cd.cell(row=2, column=startcol+1, value="value")
        r = 3
        for k, v in pairs:
            cd.cell(row=r, column=startcol, value=k)
            cd.cell(row=r, column=startcol+1, value=v); r += 1
        return startcol
    put_table(1, "Verdict", [("Toxic",len(toxic)),("Keep",len(keep)),("Monitor",len(monitor)),("Own",len(own))])
    put_table(4, "AuthorityScore", [(k, asb.get(k,0)) for k in asb_order])
    put_table(7, "Country", country.most_common(8))
    put_table(10, "TLD", tld.most_common(8))
    put_table(13, "ToxicByYear", sorted(years.items()))
    put_table(16, "ToxicIP", [(f"{k}.x", n) for k,n in ipfarm.most_common(8)])
    cd.sheet_state = "hidden"

    # ============ DASHBOARD ============
    ws = wb.create_sheet("Dashboard")
    for i,w in enumerate([3,20,16,16,16,16,16,16,16,16,16,16,16,16],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:N2")
    t = ws["B2"]; t.value = "MASTER LAWN — BACKLINK PROFILE AUDIT"
    t.font = F(20, True, WHITE); t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B3:N3")
    st = ws["B3"]; st.value = ("Client: masterlawn.com   |   Source: Semrush   |   Snapshot: 2026-09-04   |   "
                               "1,170 referring domains & 3,651 backlinks reviewed (2 QA passes)")
    st.font = F(10, False, INK_SOFT, italic=True)

    # KPI tiles row (r5)
    pct = f"{len(toxic)/total*100:.1f}%"
    tile(ws, 5, 2, "REFERRING DOMAINS", total)
    tile(ws, 5, 4, "BACKLINKS", len(bls))
    tile(ws, 5, 6, "AUTHORITY SCORE", 23)
    tile(ws, 5, 8, "TOXIC (DISAVOW)", len(toxic), valcolor=RED, fillc=VC_SOFT["TOXIC"])
    tile(ws, 5, 10, "% TOXIC", pct, valcolor=RED, fillc=VC_SOFT["TOXIC"])
    tile(ws, 5, 12, "KEEP / MONITOR", f"{len(keep)}/{len(monitor)}", valcolor=AQUA[:6], fillc=VC_SOFT["KEEP"])

    # charts
    def add_bar(anchor, title, col, ncats, color, rotate=False):
        ch = BarChart(); ch.type = "col"; ch.title = title
        ch.height = 7.2; ch.width = 12.5
        data = Reference(cd, min_col=col+1, min_row=2, max_row=2+ncats)
        cats = Reference(cd, min_col=col, min_row=3, max_row=2+ncats)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        style_barchart(ch, color)
        ws.add_chart(ch, anchor)

    # Verdict pie (per-point status colors)
    pie = PieChart(); pie.title = "Referring domains by verdict"
    pie.height = 7.2; pie.width = 12.5
    d = Reference(cd, min_col=2, min_row=2, max_row=6)
    cats = Reference(cd, min_col=1, min_row=3, max_row=6)
    pie.add_data(d, titles_from_data=True); pie.set_categories(cats)
    colors=[VC["TOXIC"],VC["KEEP"],VC["MONITOR"],VC["OWN"]]
    ser = pie.series[0]
    for i,hexc in enumerate(colors):
        dp = DataPoint(idx=i); dp.graphicalProperties.solidFill = hexc
        ser.data_points.append(dp)
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
    ws.add_chart(pie, "B9")

    add_bar("H9", "Authority Score distribution", 4, len(asb_order), ACCENT)
    add_bar("B24", "Toxic links first seen (by year)", 13, len(years), ORANGE)
    add_bar("H24", "Referring domains by country (top 8)", 7, len(country.most_common(8)), NAVY2)
    add_bar("B39", "Referring domains by TLD (top 8)", 10, len(tld.most_common(8)), VIOLET)
    add_bar("H39", "Largest toxic IP clusters (PBN farms)", 16, len(ipfarm.most_common(8)), RED)

    ws.merge_cells("B55:N58")
    note = ws["B55"]
    note.value = ("KEY FINDING — The profile is ~86% toxic: an automated directory/article/bookmark PBN and "
                  "link-selling network (single-IP farms on 64.182.x, 69.13.x, 94.46.x, 118.139.x, 159.198.75.x, "
                  "195.20.19.178), gambling/off-topic domains, and 'buy backlinks / DA-PA / telegram' anchors. "
                  f"~{pct2026:.0f}% of all referring domains first appeared in 2026 = active negative-SEO blast. Recommendation: submit the embedded "
                  "disavow (Disavow List sheet), refresh monthly, and rebuild with earned local/industry links.")
    note.font = F(10, False, INK); note.alignment = Alignment(wrap_text=True, vertical="top")
    note.fill = fill(BAND)

    # ============ EXECUTIVE SUMMARY ============
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    for i,w in enumerate([3,44,20,50],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    hrow = banner(ws, "EXECUTIVE SUMMARY — Master Lawn Backlink Audit", 4,
                  "masterlawn.com  ·  Semrush data  ·  2026-09-04")
    r = hrow + 1
    def section(title):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=title); c.font = F(12, True, NAVY)
        c.fill = fill(BAND); r += 1
    def kv(k, v, vcolor=INK):
        nonlocal r
        a = ws.cell(row=r, column=2, value=k); a.font = F(11, False, INK)
        b = ws.cell(row=r, column=3, value=v); b.font = F(11, True, vcolor)
        a.border = BORDER; b.border = BORDER; r += 1
    def bullet(txt):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value="•  " + txt); c.font = F(10, False, INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30; r += 1

    section("KEY METRICS")
    kv("Referring domains reviewed", total)
    kv("Backlinks reviewed", len(bls))
    kv("Semrush Authority Score", 23)
    kv("Organic keywords (US) / traffic per mo", "2,525 / 2,391")
    r += 1
    section("CLASSIFICATION RESULT")
    kv("TOXIC — recommended for disavow", f"{len(toxic)}  ({pct})", RED)
    kv("KEEP — legitimate / relevant", len(keep), AQUA[:6])
    kv("MONITOR — watch, do not disavow yet", len(monitor), "B7860B")
    kv("OWN — client / sister properties", len(own), ACCENT)
    r += 1
    section("KEY FINDINGS")
    bullet("~86% of referring domains are toxic — a manipulated / negative-SEO profile, not earned authority; ~90% sit at Authority Score ≤6.")
    bullet("Dominant footprint: directory/article/bookmark PBN + link-selling farms on shared IPs (64.182.x, 69.13.x, 94.46.x, 118.139.x, 159.198.75.x, 195.20.19.178).")
    bullet("Blatant spam anchors: 'buy backlinks / DA 50 PA 40 / PBN network', 'join our telegram darksidelinks', 'WhatsApp +1(226)…'; plus gambling & off-topic domains.")
    bullet(f"RECENT BLAST: ~{pct2026:.0f}% of ALL referring domains first appeared in 2026 (979 of 1,003 toxic) — "
           "consistent with an active negative-SEO attack, not slow historic decay. Disavow must be refreshed monthly & re-uploaded cumulatively. "
           "(first-seen = Semrush discovery date.)")
    bullet("A small legitimate core remains: national platforms, local news/chambers, and real lawn/landscape businesses (see KEEP).")
    r += 1
    section("RECOMMENDATION")
    bullet("Submit the embedded Disavow List (domain-level) after false-positive sweep + client sign-off; confirm ownership of masterlawninc.com / masterlawn.org / .net / midsouthturf.com / greenkingspray.com (excluded as OWN).")
    bullet("Then begin authority rebuilding (Task 4): earned local & industry links, relevance over raw DR.")

    # ============ REFERRING DOMAINS ============
    ws = wb.create_sheet("Referring Domains")
    cols = ["#","Referring domain","Verdict","Authority Score","Backlinks","IP","IP shared","Country","First seen","Last seen","Reason / evidence"]
    widths = [5,32,11,10,9,16,9,8,11,11,58]
    hrow = banner(ws, f"REFERRING DOMAINS — {total} analyzed & classified", len(cols),
                  "Verdict colour-coded. Sorted Toxic → Monitor → Keep → Own, then by toxicity.")
    header_row(ws, cols, hrow, widths)
    order = {"TOXIC":0,"MONITOR":1,"KEEP":2,"OWN":3}
    n = 0
    for rr in sorted(refs, key=lambda x:(order.get(x["verdict"],9), -x["tox"], x["domain"])):
        n += 1
        row = [n, rr["domain"], rr["verdict"], rr["ascore"], rr["backlinks"], rr["ip"],
               rr["ip_shared"], rr["country"], rr["first_seen"], rr["last_seen"], rr["reason"]]
        ws.append(row)
        rnum = ws.max_row
        vc = ws.cell(row=rnum, column=3); vc.fill = fill(VC.get(rr["verdict"],"BFBFBF"))
        vc.font = F(10, True, WHITE); vc.alignment = Alignment(horizontal="center")
        if n % 2 == 0:
            for c in range(1, len(cols)+1):
                if c != 3 and not ws.cell(row=rnum,column=c).fill.patternType:
                    ws.cell(row=rnum, column=c).fill = fill(BAND)
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(cols))}{ws.max_row}"

    # ============ ALL BACKLINKS ============
    ws = wb.create_sheet("All Backlinks")
    cols = ["#","Source URL","Target URL","Anchor text","Nofollow","Sitewide","Page AS","Referring domain","Verdict","First seen","Last seen"]
    widths = [5,46,40,42,9,8,7,26,11,11,11]
    hrow = banner(ws, f"ALL BACKLINKS — {len(bls)} links (anchor + target)", len(cols),
                  "Every backlink with its anchor text, target URL and the referring domain's verdict.")
    header_row(ws, cols, hrow, widths)
    n = 0
    for b in sorted(bls, key=lambda x:(order.get(x["verdict"],9), x["refdomain"])):
        n += 1
        ws.append([n, b["source_url"], b["target_url"], b["anchor"], b["nofollow"], b["sitewide"],
                   b["page_ascore"], b["refdomain"], b["verdict"], b["first_seen"], b["last_seen"]])
        vc = ws.cell(row=ws.max_row, column=9); vc.fill = fill(VC.get(b["verdict"],"BFBFBF"))
        vc.font = F(9, True, WHITE); vc.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(cols))}{ws.max_row}"

    # ============ DISAVOW LIST (embedded) ============
    ws = wb.create_sheet("Disavow List")
    cols = ["#","Disavow entry","Referring domain","Authority Score","Country","Reason / evidence"]
    widths = [5,34,30,10,8,60]
    hrow = banner(ws, f"DISAVOW LIST — {len(toxic)} domains (QA-PENDING)", len(cols),
                  "Domain-level entries for Google's Disavow tool. Copy column B into the .txt. "
                  "Excludes KEEP / MONITOR / OWN. Do NOT submit until sign-off.")
    header_row(ws, cols, hrow, widths)
    n = 0
    for rr in sorted(toxic, key=lambda x:(-x["tox"], x["domain"])):
        n += 1
        ws.append([n, f"domain:{rr['domain']}", rr["domain"], rr["ascore"], rr["country"], rr["reason"]])
        for c in range(1, len(cols)+1):
            if n % 2 == 0 and not ws.cell(row=ws.max_row,column=c).fill.patternType:
                ws.cell(row=ws.max_row, column=c).fill = fill(VC_SOFT["TOXIC"])
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(cols))}{ws.max_row}"

    # ============ ANCHOR ANALYSIS ============
    ws = wb.create_sheet("Anchor Analysis")
    cols = ["Anchor text","Ref domains","Backlinks","Flag"]
    hrow = banner(ws, "ANCHOR ANALYSIS", len(cols), "Spam & gambling anchors flagged.")
    header_row(ws, cols, hrow, [90,12,10,18])
    for anchor, dnum, bnum in C.ANCHORS:
        a = anchor.lower(); flag=""
        if any(t in a for t in ["buy backlinks","pbn","da 50","dofollow backlinks","telegram","whatsapp","increase your domain"]):
            flag="SPAM ANCHOR"
        elif any(t in a for t in ["casino","poker","ufa","slot"]):
            flag="GAMBLING ANCHOR"
        ws.append([anchor,dnum,bnum,flag])
        if flag:
            ws.cell(row=ws.max_row,column=4).fill = fill(VC["TOXIC"])
            ws.cell(row=ws.max_row,column=4).font = F(10, True, WHITE)

    # ============ TOXIC CLUSTERS ============
    ws = wb.create_sheet("Toxic Clusters")
    cols = ["IP /24 prefix","# toxic domains","Example domains"]
    hrow = banner(ws, "TOXIC IP CLUSTERS (PBN FARMS)", len(cols), "Multiple toxic domains sharing one IP /24 = link farm.")
    header_row(ws, cols, hrow, [18,16,95])
    ipcl = collections.defaultdict(list)
    for rr in toxic: ipcl[C.ip_prefix(rr["ip"],3)].append(rr["domain"])
    for ipf, doms in sorted(ipcl.items(), key=lambda x:-len(x[1])):
        if len(doms) >= 2:
            ws.append([ipf, len(doms), ", ".join(doms[:10])+(" ..." if len(doms)>10 else "")])

    # ============ LOCAL CITATIONS — REVIEW ============
    review_path = os.path.join(BASE, "data", "review", "local_citations.csv")
    if os.path.exists(review_path):
        ws = wb.create_sheet("Local Citations Review")
        cols = ["Referring domain","Authority Score","IP","On PBN farm?","Local links","Example local anchor","Example target page","Default recommendation"]
        widths = [30,10,16,12,10,40,46,26]
        hrow = banner(ws, "LOCAL CITATIONS — REVIEW BEFORE DISAVOW", len(cols),
                      "Toxic domains that link with a real city/service anchor. 'On farm=yes' = camouflaged spam "
                      "(safe to disavow). 'no' = standalone directory — human-review; some moved to MONITOR.")
        header_row(ws, cols, hrow, widths)
        with open(review_path, encoding="utf-8") as f:
            rd = csv.reader(f); next(rd, None)
            for row in rd:
                ws.append(row)
                onfarm = ws.cell(row=ws.max_row, column=4)
                if str(onfarm.value).lower() == "no":
                    onfarm.fill = fill(VC_SOFT["MONITOR"])
                    onfarm.font = F(10, True, "B7860B")
                else:
                    onfarm.fill = fill(VC_SOFT["TOXIC"])
        ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(cols))}{ws.max_row}"

    # ============ METHODOLOGY ============
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 110
    hrow = banner(ws, "METHODOLOGY", 2)
    lines = [
        ("Every referring domain scored on three axes → verdict, then two QA passes.", True),
        ("", False),
        ("TRUST — Semrush Authority Score; PBN/link-farm hosting (many domains on one IP /24:", True),
        ("   64.182.x, 69.13.x, 94.46.x/188.114.x, 118.139.x, 159.198.75.x, 195.20.19.178).", False),
        ("RELEVANCE — off-topic high-risk verticals (gambling/adult/pharma/essay) = toxic;", True),
        ("   link-scheme & directory/article/bookmark names = toxic; on-topic lawn/landscape or", False),
        ("   local (Memphis/Germantown/Bartlett/Olive Branch/Huntsville) + real authority = keep.", False),
        ("LANGUAGE/GEO — foreign hosting (SE/SG/MD/VN/BD) with English commercial anchors to a", True),
        ("   hyper-local US lawn business = unnatural = toxic signal.", False),
        ("", False),
        ("VERDICTS — TOXIC (disavow) · MONITOR (watch) · KEEP (legit) · OWN (client, never disavow).", True),
        ("QA — Pass 1 adjudicated borderline + verified keeps; Pass 2 (adversarial, 8 agents)", True),
        ("   re-checked all toxic to rescue false-positives and re-verified keeps for false-negatives.", False),
        ("", False),
        ("Disavow is conservative: only TOXIC entries; MONITOR/KEEP/OWN excluded; client sign-off required.", True),
    ]
    rr2 = hrow
    for txt, bold in lines:
        c = ws.cell(row=rr2, column=2, value=txt); c.font = F(10, bold, NAVY if bold else INK)
        rr2 += 1

    out = os.path.join(BASE, "MasterLawn_Backlink_Audit.xlsx")
    # order sheets: Dashboard, Exec, Ref Domains, All Backlinks, Disavow, Anchor, Clusters, Method, (hidden ChartData)
    wb.move_sheet("Dashboard", -(wb.sheetnames.index("Dashboard")))
    desired = ["Dashboard","Executive Summary","Referring Domains","All Backlinks","Disavow List",
               "Local Citations Review","Anchor Analysis","Toxic Clusters","Methodology","Chart Data"]
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)
    wb.active = 0
    wb.save(out)
    print("saved", out)
    print(f"verdicts: TOXIC {len(toxic)} | KEEP {len(keep)} | MONITOR {len(monitor)} | OWN {len(own)}")

if __name__ == "__main__":
    build()
