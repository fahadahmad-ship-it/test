#!/usr/bin/env python3
"""
Master Lawn — backlink profile classifier + Excel deliverable builder.

Reads raw Semrush exports:
  seo/data/refdomains_raw.csv   (domain;domain_ascore;backlinks_num;ip;country;first_seen;last_seen)
  seo/data/backlinks_raw.csv    (source_url;target_url;anchor;nofollow;sitewide;page_ascore;first_seen;last_seen)

Classifies EVERY referring domain on three axes (Trust, Relevance, Language)
into a verdict (KEEP / MONITOR / TOXIC / OWN) with machine-readable reason codes,
then writes three workbooks:
  seo/MasterLawn_Backlink_Audit.xlsx
  seo/MasterLawn_Disavow.xlsx  (+ seo/disavow_masterlawn.txt)
  seo/MasterLawn_Baseline_Benchmark.xlsx
"""
import csv, os, re, datetime, collections

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")

# ---------------------------------------------------------------- helpers
def epoch_to_date(v):
    try:
        return datetime.datetime.utcfromtimestamp(int(float(v))).strftime("%Y-%m-%d")
    except Exception:
        return ""

def root_domain(host):
    host = host.strip().lower()
    host = re.sub(r"^https?://", "", host)
    host = host.split("/")[0].split("?")[0]
    if host.startswith("www."):
        host = host[4:]
    return host

def ip_prefix(ip, n=3):
    parts = (ip or "").split(".")
    return ".".join(parts[:n]) if len(parts) >= n else (ip or "")

# ---------------------------------------------------------------- rule data
# Client-owned / sister brand properties — never disavow these.
OWN = {
    "masterlawn.com", "masterlawn.org", "masterlawn.net", "masterlawninc.com",
    "midsouthturf.com", "greenkingspray.com",
}

# Known PBN / link-farm hosting clusters observed in this profile (IP /24 prefixes).
PBN_IP_PREFIXES = {
    "64.182", "69.13",                      # directory/article PBN farm (US)
    "94.46", "185.39", "185.41", "185.51", "188.114", "185.196", "185.198",  # Sweden/EU directory farm
    "118.139",                              # Singapore farm
    "184.168.111", "184.168.113", "184.168.115", "184.168.116",  # SG stat/whois spam
    "15.204.130", "15.204.135",             # bookmark spam
    "159.198.75",                           # SEO/link-shop PBN
    "195.20.19",                            # Moldova link-shortener/spam network
}

# High-risk off-topic verticals (word tokens).
VERTICAL_TOKENS = [
    "casino", "poker", "ufa", "slot", "bookie", "bookies", "gambl", "betting",
    "porn", "xxx", "adult", "escort", "pharma", "viagra", "cialis", "crypto",
    "essay", "essays",
]

# Link-scheme / SEO / PBN name tokens.
SEO_TOKENS = [
    "seo", "backlink", "dofollow", "pbn", "guestpost", "linkbuild", "buybacklink",
    "rankvance", "linkseo", "seolink", "ranker", "webranks", "worldwideranks",
    "multiranks", "softranks", "globalrank", "ranking1m", "99ranks", "99w",
]

# Directory / article / bookmark spam name tokens.
DIR_TOKENS = [
    "directory", "directori", "listing", "listingz", "weblink", "weblinx",
    "webdir", "bizdir", "bizlisting", "bizlistings", "submitweb", "submitarticle",
    "submitbest", "bookmark", "article", "articles", "weblist", "webindex",
    "web-directory", "webdirectory", "yellowpage", "linkodirectory", "dirsearch",
]

# Spammy / low-trust TLDs.
SPAM_TLDS = {
    "icu", "cfd", "sbs", "top", "monster", "party", "website", "space", "store",
    "click", "shop", "cv", "im", "ws", "lc", "bz", "id", "fyi", "art", "help",
    "homes", "pro", "world", "online", "live",
}

# Foreign geos that should not naturally link to a Memphis/N-MS/Huntsville lawn co.
FOREIGN_SPAM_COUNTRIES = {"se", "sg", "md", "vn", "bd", "cz", "id", "in", "ro", "pk"}

# Known-legitimate large platforms / aggregators (keep; some are neutral scrapers).
KNOWN_GOOD = {
    "yellowpages.com", "superpages.com", "dexknows.com", "citysquares.com",
    "agreatertown.com", "hub.biz", "owler.com", "birdeye.com", "pitchbook.com",
    "rocketreach.co", "theorg.com", "devpost.com", "yourgreenpal.com",
    "housedigest.com", "bing.com", "yahoo.com", "alibaba.com", "dhgate.com",
    "express.co.uk", "thesun.ie", "extension.org", "enigma.com", "strollmag.com",
    "landscapeprofessionals.org", "totallandscapecare.com", "landscapeleadership.com",
    "neustarlocaleze.biz", "growthzoneapp.com", "regionaldirectory.us",
    "yplocal.us", "find-us-here.com", "loginslink.com", "sitelike.org",
    "natureworldnews.com", "olivebranchms.com", "olivebranchmagazine.com",
    "desotocountynews.com", "southavenchamber.com", "hsvchamber.org",
    "bestprosintown.com", "mymilitarybenefits.com", "propartsdirect.net",
}

# Relevance: on-topic tokens (lawn / landscape / local home services / geo).
RELEVANT_TOKENS = [
    "lawn", "landscap", "turf", "garden", "grass", "yard", "mosquito", "irrigation",
    "greenhouse", "plant", "nursery", "bloom", "flower", "gard", "outdoor",
    "chamber", "contractor", "gardening",
]
GEO_TOKENS = ["memphis", "germantown", "collierville", "bartlett", "olivebranch",
              "olive-branch", "desoto", "huntsville", "southaven", "tennessee",
              "mississippi", "hsv"]

def has_token(name, tokens):
    return any(t in name for t in tokens)

# ---------------------------------------------------------------- classifier
def classify(domain, ascore, backlinks, ip, country):
    name = domain.lower()
    tld = name.rsplit(".", 1)[-1] if "." in name else ""
    ipf3 = ip_prefix(ip, 3)
    ipf2 = ip_prefix(ip, 2)
    reasons = []
    tox = 0  # toxicity score

    # ---- OWN properties (never disavow) ----
    if name in OWN:
        return ("OWN", 0, "own-property", ["client-owned/sister-brand property"])

    if name in KNOWN_GOOD:
        # still record why it's kept
        rel = "relevant" if has_token(name, RELEVANT_TOKENS + GEO_TOKENS) else "neutral-authority"
        return ("KEEP", 0, "known-good", [f"recognized legitimate platform ({rel})"])

    # ---- Trust axis ----
    if ipf3 in PBN_IP_PREFIXES or ipf2 in PBN_IP_PREFIXES:
        tox += 3; reasons.append(f"pbn-ip-cluster:{ipf3}")
    if ascore <= 6:
        tox += 1; reasons.append(f"very-low-authority(AS{ascore})")

    # ---- Relevance axis: spam name patterns ----
    if has_token(name, VERTICAL_TOKENS):
        tox += 4; reasons.append("high-risk-vertical(gambling/adult/pharma/essay)")
    if has_token(name, SEO_TOKENS):
        tox += 3; reasons.append("seo/link-scheme-name")
    if has_token(name, DIR_TOKENS):
        tox += 3; reasons.append("directory/article/bookmark-spam-name")

    # ---- TLD signal ----
    if tld in SPAM_TLDS:
        tox += 1; reasons.append(f"spam-tld(.{tld})")

    # ---- Language / geo signal ----
    if country in FOREIGN_SPAM_COUNTRIES:
        tox += 1; reasons.append(f"foreign-geo-cluster({country})")

    # ---- Relevance offset (genuinely on-topic + real authority) ----
    relevant = has_token(name, RELEVANT_TOKENS) or has_token(name, GEO_TOKENS)
    if relevant and ascore >= 15 and tox < 3:
        return ("KEEP", tox, "relevant-authority",
                ["on-topic/local + real authority; no strong spam signal"])

    # ---- Verdict thresholds ----
    if tox >= 3:
        return ("TOXIC", tox, reasons[0] if reasons else "spam-composite", reasons)
    if ascore >= 20 and tox == 0:
        return ("KEEP", tox, "authority-no-spam",
                ["AS>=20 with no spam signal"])
    if tox == 0 and relevant:
        return ("MONITOR", tox, "relevant-lowauth",
                ["on-topic but low authority; monitor, do not disavow"])
    # low authority, weak-but-present signal, or unknown → manual review / monitor
    if tox >= 1:
        return ("MONITOR", tox, reasons[0], reasons + ["weak signal — human QA"])
    return ("MONITOR", tox, "unclassified", ["no strong signal — human QA"])

# ---------------------------------------------------------------- load data
def load_refdomains():
    path = os.path.join(DATA, "refdomains_raw.csv")
    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=";")
        header = next(reader)
        for r in reader:
            if not r or not r[0] or r[0] == "domain":
                continue
            # tolerant positional parse
            domain = r[0].strip()
            try: ascore = int(float(r[1]))
            except Exception: ascore = 0
            try: backlinks = int(float(r[2]))
            except Exception: backlinks = 0
            ip = r[3].strip() if len(r) > 3 else ""
            country = (r[4].strip().lower() if len(r) > 4 else "")
            first = r[5] if len(r) > 5 else ""
            last = r[6] if len(r) > 6 else ""
            rows.append(dict(domain=domain, ascore=ascore, backlinks=backlinks,
                             ip=ip, country=country,
                             first_seen=epoch_to_date(first),
                             last_seen=epoch_to_date(last)))
    return rows

def load_backlinks():
    path = os.path.join(DATA, "backlinks_raw.csv")
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, encoding="utf-8", errors="replace") as f:
        first_line = f.readline()  # header
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split(";")
            if len(parts) < 8:
                # anchor may itself contain ';' — but we split fixed head(2)+tail(5)
                if len(parts) < 7:
                    continue
            source_url = parts[0]
            target_url = parts[1]
            # last 5 are fixed: nofollow;sitewide;page_ascore;first_seen;last_seen
            tail = parts[-5:]
            anchor = ";".join(parts[2:-5]) if len(parts) > 7 else parts[2]
            nofollow, sitewide, pasc, fs, ls = tail
            rows.append(dict(
                source_url=source_url, target_url=target_url, anchor=anchor,
                nofollow=nofollow, sitewide=sitewide, page_ascore=pasc,
                first_seen=epoch_to_date(fs), last_seen=epoch_to_date(ls),
                refdomain=root_domain(source_url)))
    return rows

# ---------------------------------------------------------------- build excel
def build():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    refs = load_refdomains()
    bls = load_backlinks()

    # load QA overrides (agent-reviewed verdicts), if present
    overrides = {}
    override_files = [os.path.join(DATA, "qa", fn) for fn in
                     ("monitor_A_reviewed.csv", "monitor_B_reviewed.csv", "keep_reviewed.csv")]
    # qa2 = critical second-pass review; loaded last so it takes precedence
    override_files += [os.path.join(DATA, "qa2", f"toxic_{i}_reviewed.csv") for i in range(1, 8)]
    override_files.append(os.path.join(DATA, "qa2", "keep_final_reviewed.csv"))
    for p in override_files:
        if not os.path.exists(p):
            continue
        with open(p, encoding="utf-8", errors="replace") as f:
            rd = csv.reader(f)
            hdr = next(rd, None)
            for row in rd:
                if len(row) < 2 or not row[0].strip():
                    continue
                dom = root_domain(row[0])
                fv = row[1].strip().upper()
                note = row[3].strip() if len(row) > 3 else ""
                if fv in ("TOXIC", "MONITOR", "KEEP"):
                    overrides[dom] = (fv, note)

    # classify referring domains
    verdict_counts = collections.Counter()
    reason_counts = collections.Counter()
    for r in refs:
        v, score, code, reasons = classify(r["domain"], r["ascore"], r["backlinks"],
                                            r["ip"], r["country"])
        ov = overrides.get(r["domain"].lower())
        if ov and v not in ("OWN",):
            if ov[0] != v:
                code = "qa-override:" + code
                reasons.append(f"QA agent → {ov[0]}" + (f" ({ov[1]})" if ov[1] else ""))
            v = ov[0]
        r["verdict"] = v
        r["tox_score"] = score
        r["reason_code"] = code
        r["reasons"] = "; ".join(reasons)
        verdict_counts[v] += 1
        reason_counts[code] += 1

    # map verdict back to each backlink via its refdomain
    ref_by_domain = {r["domain"]: r for r in refs}
    for b in bls:
        rd = b["refdomain"]
        rr = ref_by_domain.get(rd)
        b["verdict"] = rr["verdict"] if rr else "UNMAPPED"
        b["ref_ascore"] = rr["ascore"] if rr else ""
        b["reason_code"] = rr["reason_code"] if rr else ""

    toxic = [r for r in refs if r["verdict"] == "TOXIC"]
    monitor = [r for r in refs if r["verdict"] == "MONITOR"]
    keep = [r for r in refs if r["verdict"] == "KEEP"]
    own = [r for r in refs if r["verdict"] == "OWN"]

    # ---------- styles ----------
    HDR = Font(bold=True, color="FFFFFF", size=11)
    HFILL = PatternFill("solid", fgColor="1F4E78")
    TITLE = Font(bold=True, size=16, color="1F4E78")
    SUB = Font(italic=True, size=10, color="555555")
    BOLD = Font(bold=True)
    VFILL = {
        "TOXIC": PatternFill("solid", fgColor="F4CCCC"),
        "MONITOR": PatternFill("solid", fgColor="FFF2CC"),
        "KEEP": PatternFill("solid", fgColor="D9EAD3"),
        "OWN": PatternFill("solid", fgColor="D0E0E3"),
        "UNMAPPED": PatternFill("solid", fgColor="EEEEEE"),
    }
    thin = Side(style="thin", color="DDDDDD")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HDR; cell.fill = HFILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def autowidth(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ================================================= WORKBOOK 1: AUDIT
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Master Lawn — Backlink Profile Audit"; ws["A1"].font = TITLE
    ws["A2"] = "Client: https://www.masterlawn.com/   |   Data source: Semrush   |   Snapshot: 2026-09-04"
    ws["A2"].font = SUB
    total = len(refs)
    rows = [
        ["", ""],
        ["AUDIT SCOPE", ""],
        ["Referring domains reviewed", total],
        ["Backlinks reviewed", len(bls)],
        ["", ""],
        ["CLASSIFICATION RESULT (referring domains)", ""],
        ["TOXIC — recommend disavow", len(toxic)],
        ["MONITOR — human QA / watch (do NOT disavow yet)", len(monitor)],
        ["KEEP — legitimate / relevant authority", len(keep)],
        ["OWN — client-owned / sister brand", len(own)],
        ["", ""],
        ["% of referring domains toxic", f"{len(toxic)/total*100:.1f}%"],
        ["% of referring domains at Authority Score <=6", ""],  # filled below
    ]
    lowauth = sum(1 for r in refs if r["ascore"] <= 6)
    rows[-1][1] = f"{lowauth/total*100:.1f}%"
    r0 = 4
    for i, (a, b) in enumerate(rows):
        ws.cell(row=r0 + i, column=1, value=a)
        ws.cell(row=r0 + i, column=2, value=b)
        if a and a.isupper() and b == "":
            ws.cell(row=r0 + i, column=1).font = BOLD
    for lbl in ("TOXIC — recommend disavow",):
        pass
    # color the verdict rows
    color_map = {"TOXIC — recommend disavow": "F4CCCC",
                 "MONITOR — human QA / watch (do NOT disavow yet)": "FFF2CC",
                 "KEEP — legitimate / relevant authority": "D9EAD3",
                 "OWN — client-owned / sister brand": "D0E0E3"}
    for i, (a, b) in enumerate(rows):
        if a in color_map:
            for c in (1, 2):
                ws.cell(row=r0 + i, column=c).fill = PatternFill("solid", fgColor=color_map[a])
                ws.cell(row=r0 + i, column=c).font = BOLD
    autowidth(ws, [52, 16])

    note_row = r0 + len(rows) + 2
    ws.cell(row=note_row, column=1,
            value="Key finding: the profile is dominated by an automated directory/article/PBN "
                  "link-spam footprint (single-IP farms, spammy TLDs, foreign hosting clusters, "
                  "'buy backlinks / DA-PA / telegram' anchors). This is manipulative/negative-SEO "
                  "contamination, not earned authority — hence the critical audit + disavow.")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 3, end_column=2)

    # ---- Referring Domains (classified) ----
    ws = wb.create_sheet("Referring Domains")
    cols = ["Referring domain", "Verdict", "Toxicity score", "Authority Score",
            "Backlinks", "IP", "Country", "First seen", "Last seen",
            "Reason code", "Evidence / reasons"]
    ws.append(cols); style_header(ws, len(cols))
    order = {"TOXIC": 0, "MONITOR": 1, "KEEP": 2, "OWN": 3}
    for r in sorted(refs, key=lambda x: (order.get(x["verdict"], 9), -x["tox_score"], x["domain"])):
        ws.append([r["domain"], r["verdict"], r["tox_score"], r["ascore"], r["backlinks"],
                   r["ip"], r["country"], r["first_seen"], r["last_seen"],
                   r["reason_code"], r["reasons"]])
        ws.cell(row=ws.max_row, column=2).fill = VFILL.get(r["verdict"], VFILL["UNMAPPED"])
    autowidth(ws, [34, 10, 9, 9, 9, 16, 8, 11, 11, 26, 60])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ---- All Backlinks (classified) ----
    ws = wb.create_sheet("All Backlinks")
    cols = ["Source URL", "Target URL", "Anchor text", "Nofollow", "Sitewide",
            "Page AS", "Referring domain", "Domain verdict", "First seen", "Last seen"]
    ws.append(cols); style_header(ws, len(cols))
    for b in sorted(bls, key=lambda x: (order.get(x["verdict"], 9), x["refdomain"])):
        ws.append([b["source_url"], b["target_url"], b["anchor"], b["nofollow"],
                   b["sitewide"], b["page_ascore"], b["refdomain"], b["verdict"],
                   b["first_seen"], b["last_seen"]])
        ws.cell(row=ws.max_row, column=8).fill = VFILL.get(b["verdict"], VFILL["UNMAPPED"])
    autowidth(ws, [46, 40, 40, 9, 8, 7, 26, 12, 11, 11])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ---- Anchor analysis ----
    ws = wb.create_sheet("Anchor Analysis")
    ws.append(["Anchor text", "Ref domains", "Backlinks", "Flag"])
    style_header(ws, 4)
    for anchor, dnum, bnum in ANCHORS:
        flag = ""
        a = anchor.lower()
        if any(t in a for t in ["buy backlinks", "pbn", "da 50", "dofollow backlinks",
                                "telegram", "whatsapp", "increase your domain"]):
            flag = "SPAM ANCHOR"
        elif any(t in a for t in ["casino", "poker", "ufa", "slot"]):
            flag = "GAMBLING ANCHOR"
        ws.append([anchor, dnum, bnum, flag])
        if flag:
            ws.cell(row=ws.max_row, column=4).fill = VFILL["TOXIC"]
    autowidth(ws, [90, 12, 10, 16])

    # ---- Distributions ----
    ws = wb.create_sheet("Distributions")
    ws["A1"] = "Authority Score distribution (referring domains)"; ws["A1"].font = BOLD
    ws.append([]); ws.append(["Authority Score", "Ref domains"])
    style_header(ws, 2, row=3)
    for asc, n in ASCORE:
        ws.append([asc, n])
    start = ws.max_row + 3
    ws.cell(row=start, column=1, value="Referring domains by TLD zone").font = BOLD
    ws.cell(row=start + 1, column=1, value="TLD"); ws.cell(row=start + 1, column=2, value="Ref domains")
    ws.cell(row=start + 1, column=3, value="Backlinks")
    for c in range(1, 4):
        ws.cell(row=start + 1, column=c).font = HDR; ws.cell(row=start + 1, column=c).fill = HFILL
    rr = start + 2
    for zone, d, b in TLD:
        ws.cell(row=rr, column=1, value=zone); ws.cell(row=rr, column=2, value=d)
        ws.cell(row=rr, column=3, value=b); rr += 1
    start2 = rr + 2
    ws.cell(row=start2, column=1, value="Referring domains by country (IP geo)").font = BOLD
    ws.cell(row=start2 + 1, column=1, value="Country"); ws.cell(row=start2 + 1, column=2, value="Ref domains")
    ws.cell(row=start2 + 1, column=3, value="Backlinks")
    for c in range(1, 4):
        ws.cell(row=start2 + 1, column=c).font = HDR; ws.cell(row=start2 + 1, column=c).fill = HFILL
    rr = start2 + 2
    for country, d, b in GEO:
        ws.cell(row=rr, column=1, value=country); ws.cell(row=rr, column=2, value=d)
        ws.cell(row=rr, column=3, value=b); rr += 1
    autowidth(ws, [26, 14, 12])

    # ---- Toxic clusters (by IP) ----
    ws = wb.create_sheet("Toxic Clusters")
    ws.append(["IP /24 prefix", "# toxic domains on this IP", "Example domains"])
    style_header(ws, 3)
    ipcl = collections.defaultdict(list)
    for r in toxic:
        ipcl[ip_prefix(r["ip"], 3)].append(r["domain"])
    for ipf, doms in sorted(ipcl.items(), key=lambda x: -len(x[1])):
        if len(doms) >= 2:
            ws.append([ipf, len(doms), ", ".join(doms[:8]) + (" ..." if len(doms) > 8 else "")])
    autowidth(ws, [18, 26, 90])

    # ---- Methodology ----
    ws = wb.create_sheet("Methodology")
    method = [
        ("Master Lawn — Backlink Audit Methodology", TITLE),
        ("", None),
        ("Every referring domain was scored on three axes and assigned a verdict.", BOLD),
        ("", None),
        ("TRUST", BOLD),
        ("  • Semrush Authority Score (domain).  AS<=6 with a spam signal = red flag.", None),
        ("  • PBN/link-farm hosting: multiple domains sharing a single IP /24 (e.g. 64.182.x,", None),
        ("    69.13.x directory farm; 94.46.x / 188.114.x Sweden farm; 118.139.x Singapore;", None),
        ("    159.198.75.x SEO link-shop; 195.20.19.178 Moldova link-shortener network).", None),
        ("", None),
        ("RELEVANCE", BOLD),
        ("  • Off-topic high-risk verticals (gambling/adult/pharma/essay) = toxic.", None),
        ("  • Link-scheme names (seo, backlink, dofollow, pbn, rank...) = toxic.", None),
        ("  • Directory/article/bookmark spam names (directory, listing, bookmark, article...) = toxic.", None),
        ("  • On-topic (lawn/landscape/turf/garden/mosquito) or local (Memphis/Germantown/", None),
        ("    Bartlett/Olive Branch/Huntsville) + real authority = KEEP.", None),
        ("", None),
        ("LANGUAGE / GEO", BOLD),
        ("  • Foreign hosting clusters (SE, SG, MD, VN, BD...) linking with English commercial", None),
        ("    lawn-care anchors to a hyper-local US service business = unnatural = toxic signal.", None),
        ("", None),
        ("VERDICTS", BOLD),
        ("  • TOXIC   → composite toxicity score >=3 → disavow candidate.", None),
        ("  • MONITOR → weak/ambiguous signal → human QA, do NOT disavow yet.", None),
        ("  • KEEP    → legitimate, relevant, or real-authority with no spam signal.", None),
        ("  • OWN     → client-owned / sister-brand property (never disavow).", None),
        ("", None),
        ("NOTE: Disavow is conservative. Only TOXIC domains are proposed for the disavow file;", None),
        ("MONITOR and KEEP are excluded. Final disavow requires human + client sign-off.", None),
    ]
    for i, (txt, fnt) in enumerate(method, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if fnt: c.font = fnt
    autowidth(ws, [100])

    wb.save(os.path.join(BASE, "MasterLawn_Backlink_Audit.xlsx"))

    # ================================================= WORKBOOK 2: DISAVOW
    wb2 = Workbook()
    ws = wb2.active; ws.title = "Disavow (domain-level)"
    ws["A1"] = "Master Lawn — Disavow Candidate List (QA-PENDING)"; ws["A1"].font = TITLE
    ws["A2"] = ("All TOXIC referring domains. Review before submission. Disavow entries are "
                "domain-level (domain:<root>). MONITOR/KEEP/OWN are intentionally excluded.")
    ws["A2"].font = SUB
    ws.append([]); ws.append([])
    hdr = ["disavow entry", "referring domain", "Authority Score", "IP", "Country",
           "reason code", "evidence"]
    ws.append(hdr); style_header(ws, len(hdr), row=5)
    for r in sorted(toxic, key=lambda x: (-x["tox_score"], x["domain"])):
        ws.append([f"domain:{r['domain']}", r["domain"], r["ascore"], r["ip"],
                   r["country"], r["reason_code"], r["reasons"]])
    autowidth(ws, [34, 30, 9, 16, 8, 26, 60])
    ws.auto_filter.ref = f"A5:{get_column_letter(len(hdr))}{ws.max_row}"

    ws2 = wb2.create_sheet("README")
    readme = [
        "MASTER LAWN — DISAVOW FILE PREPARATION",
        "",
        "STATUS: QA-PENDING. This is the candidate list produced by the automated + agent-",
        "reviewed audit. Do NOT upload until the false-positive sweep and two-analyst +",
        "client sign-off are complete.",
        "",
        "WHAT IS INCLUDED: every referring domain classified TOXIC (composite score >=3):",
        "  - PBN / directory / article / bookmark link-farms (single-IP clusters)",
        "  - SEO / link-selling / 'buy backlinks' / rank-scheme domains",
        "  - gambling / adult / off-topic high-risk verticals",
        "  - foreign-hosting spam clusters linking with English lawn-care anchors",
        "",
        "WHAT IS EXCLUDED (deliberately):",
        "  - OWN client/sister properties (masterlawn.org/.net, masterlawninc.com,",
        "    midsouthturf.com, greenkingspray.com)",
        "  - KEEP: legitimate directories/platforms, local news, chambers, industry sites",
        "  - MONITOR: low-authority but not clearly harmful — watch, do not disavow",
        "",
        "SUBMISSION:",
        "  1. Export the domain: lines (see disavow_masterlawn.txt).",
        "  2. Validate format (UTF-8, one entry per line, domain: syntax, comments with #).",
        "  3. False-positive sweep vs KEEP/MONITOR + Google Search Console 'Top linking sites'.",
        "  4. Two-analyst + client sign-off.",
        "  5. Upload to the correct GSC property via the Disavow Links tool.",
        "  6. The file is CUMULATIVE — always re-upload the full list. Refresh monthly",
        "     because the spam injection is ongoing.",
    ]
    for i, t in enumerate(readme, 1):
        c = ws2.cell(row=i, column=1, value=t)
        if i == 1: c.font = TITLE
        elif t.endswith(":") or t.isupper(): c.font = BOLD
    autowidth(ws2, [95])
    wb2.save(os.path.join(BASE, "MasterLawn_Disavow.xlsx"))

    # plain-text disavow file (Google format)
    with open(os.path.join(BASE, "disavow_masterlawn.txt"), "w", encoding="utf-8") as f:
        f.write("# Master Lawn — disavow file (QA-PENDING, do not submit until signed off)\n")
        f.write(f"# Generated {datetime.date.today()} from Semrush backlink audit.\n")
        f.write(f"# {len(toxic)} toxic referring domains (domain-level).\n#\n")
        for r in sorted(toxic, key=lambda x: x["domain"]):
            f.write(f"domain:{r['domain']}\n")

    # ================================================= WORKBOOK 3: BASELINE
    wb3 = Workbook()
    ws = wb3.active; ws.title = "Baseline Benchmark"
    ws["A1"] = "Master Lawn — Critical SEO Baseline Benchmark"; ws["A1"].font = TITLE
    ws["A2"] = "Source: Semrush  |  Snapshot date: 2026-09-04  |  Frozen 'before' state for measuring disavow + rebuild impact"
    ws["A2"].font = SUB
    blocks = [
        ("AUTHORITY & BACKLINK PROFILE (Semrush)", [
            ("Authority Score", 23, "Low; inflated by spam volume, not earned trust"),
            ("Total backlinks", 3613, ""),
            ("Referring domains", 1167, "≈90% at Authority Score <=6 (spam/PBN bulk)"),
            ("Referring IPs", 1151, ""),
            ("Referring IPs (Class C)", 792, ""),
            ("Follow backlinks", 2877, ""),
            ("Nofollow backlinks", 743, ""),
            ("Sponsored / UGC", "0 / 5", ""),
            ("Text / Image links", "3243 / 104", ""),
        ]),
        ("TOXICITY (this audit)", [
            ("Referring domains classified TOXIC", len(toxic), "recommend disavow"),
            ("Referring domains MONITOR", len(monitor), "human QA"),
            ("Referring domains KEEP", len(keep), "legitimate/relevant"),
            ("Referring domains OWN", len(own), "client/sister properties"),
            ("% toxic", f"{len(toxic)/len(refs)*100:.1f}%", ""),
        ]),
        ("ORGANIC VISIBILITY (Semrush, US)", [
            ("Semrush Rank", 595080, ""),
            ("Organic keywords", 2525, ""),
            ("Keywords in positions 1-3", 104, ""),
            ("Keywords in positions 4-10", 235, ""),
            ("Keywords in positions 11-20", 223, ""),
            ("Organic traffic / month", 2391, ""),
            ("Organic traffic value / month (USD)", 10324, ""),
            ("Paid keywords / traffic", "0 / 0", ""),
        ]),
        ("PENDING CLIENT ACCESS (source of truth once granted)", [
            ("Google Search Console", "clicks / impressions / CTR / avg position", "request access"),
            ("GA4", "organic sessions & conversions", "request access"),
            ("Rank tracker (local pack)", "to configure", "Memphis/Germantown/Bartlett/Olive Branch/Huntsville"),
            ("Confirm ownership", "masterlawninc.com / masterlawn.org / .net / midsouthturf / greenkingspray", ""),
        ]),
    ]
    rr = 4
    for title, items in blocks:
        ws.cell(row=rr, column=1, value=title).font = BOLD
        ws.cell(row=rr, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=rr, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=rr, column=3).fill = PatternFill("solid", fgColor="D9E1F2")
        rr += 1
        for name, val, note in items:
            ws.cell(row=rr, column=1, value=name)
            ws.cell(row=rr, column=2, value=val)
            ws.cell(row=rr, column=3, value=note)
            rr += 1
        rr += 1
    autowidth(ws, [42, 22, 58])

    # local keyword tracking sheet
    ws = wb3.create_sheet("Local Keyword Set")
    ws.append(["Service", "Geo modifiers to track"])
    style_header(ws, 2)
    services = ["lawn care", "lawn fertilization", "weed control", "lawn aeration",
                "mosquito control", "lawn treatment", "lawn maintenance"]
    geos = "Memphis TN; Germantown TN; Collierville TN; Bartlett TN; Olive Branch MS; Southaven MS; Huntsville AL"
    for s in services:
        ws.append([s, geos])
    autowidth(ws, [22, 80])

    # change-log sheet
    ws = wb3.create_sheet("Change Log")
    ws.append(["Date", "Authority Score", "Ref domains", "Toxic domains", "Organic KW",
               "Organic traffic/mo", "Note"])
    style_header(ws, 7)
    ws.append(["2026-09-04", 23, 1167, len(toxic), 2525, 2391, "Baseline captured (pre-disavow)"])
    autowidth(ws, [12, 14, 12, 13, 11, 17, 40])
    wb3.save(os.path.join(BASE, "MasterLawn_Baseline_Benchmark.xlsx"))

    # ---------------------------------------------------------------- console
    print("=== CLASSIFICATION SUMMARY ===")
    print(f"Referring domains reviewed : {len(refs)}")
    print(f"Backlinks reviewed         : {len(bls)}")
    print(f"  TOXIC   : {len(toxic)}")
    print(f"  MONITOR : {len(monitor)}")
    print(f"  KEEP    : {len(keep)}")
    print(f"  OWN     : {len(own)}")
    print(f"  toxic % : {len(toxic)/len(refs)*100:.1f}%")
    print("Top reason codes:")
    for code, n in reason_counts.most_common(12):
        print(f"   {n:5d}  {code}")
    print("Files written: MasterLawn_Backlink_Audit.xlsx, MasterLawn_Disavow.xlsx, "
          "MasterLawn_Baseline_Benchmark.xlsx, disavow_masterlawn.txt")


# ---------------------------------------------------------------- embedded aggregates (Semrush, 2026-09-04)
ASCORE = [(0,12),(2,876),(3,23),(4,33),(5,58),(6,43),(7,7),(8,4),(9,2),(10,3),(11,4),
    (12,7),(13,9),(14,3),(15,3),(16,6),(17,1),(18,3),(19,2),(21,6),(22,5),(23,4),(24,3),
    (25,1),(26,3),(27,4),(28,3),(29,2),(30,2),(31,4),(32,4),(33,2),(34,1),(35,3),(37,1),
    (38,1),(41,1),(42,2),(44,1),(45,2),(46,1),(47,1),(49,1),(51,1),(52,1),(58,1),(59,1),
    (63,1),(68,2),(69,1),(70,1),(79,1),(84,1),(96,1),(100,1)]
TLD = [("com",625,2013),("org",145,348),("net",112,263),("co",59,128),("biz",41,141),
    ("info",33,87),("us",27,56),("in",14,68),("dev",10,14),("uk",10,17),("top",6,28),
    ("pro",6,44),("online",5,11),("space",4,18),("blog",3,5),("shop",3,16),("au",3,11),
    ("art",3,14),("eu",3,17),("sbs",3,63),("cv",3,20),("io",3,8),("website",2,26),
    ("monster",2,27),("ai",2,3),("help",2,3),("ca",2,3),("store",2,31),("lk",2,3),
    ("pe",1,1),("party",1,4),("cfd",1,25),("pa",1,1),("world",1,7),("ro",1,1),("de",1,2),
    ("show",1,2),("pk",1,1),("cz",1,1),("za",1,1)]
GEO = [("United States",765,2014),("Sweden",116,155),("Singapore",72,556),("Moldova",19,102),
    ("France",12,74),("United Kingdom",8,16),("The Netherlands",4,16),("India",4,8),
    ("Spain",2,2),("Germany",2,2),("Vietnam",1,1),("Finland",1,2),("Bangladesh",1,1),
    ("Australia",1,1),("Switzerland",1,1),("Canada",1,1),("Czechia",1,1)]
ANCHORS = [("masterlawn.com",318,401),("<EmptyAnchor>",11,314),("visit website",170,212),
    ("masterlawn.org",83,204),("masterlawninc.com",94,168),("greenkingspray.com",94,142),
    ("midsouthturf.com",87,141),("masterlawn.net",44,90),
    ("www.masterlawn.com/lawn-care-and-spray-huntsville-al",3,63),
    ("icon - external link visit website",1,62),("website",16,51),("masterlawn",40,41),
    ("high quality dofollow backlinks da 50 pa 40 premium pbn network service masterlawn.com rank first page google fast seo link building buy backlinks online cheap",26,29),
    ("high quality dofollow backlinks da 50 pa 40 premium pbn network service masterlawn.org rank first page google fast seo link building buy backlinks online cheap",26,29),
    ("high quality dofollow backlinks da 50 pa 40 premium pbn network service masterlawninc.com rank first page google fast seo link building buy backlinks online cheap",26,29),
    ("high quality dofollow backlinks da 50 pa 40 premium pbn network service midsouthturf.com rank first page google fast seo link building buy backlinks online cheap",26,29),
    ("mosquito control huntsville al",24,28),("weed control company huntsville al",22,28),
    ("lawn care germantown tn",26,26),("join our telegram https://t.me/s/darksidelinks",18,23),
    ("lawn treatment services huntsville al",18,21),
    ("i will increase your domain authority/rating from 0 -70 in 5 days. whatsapp us on: +1(226)799-7565. thank you.",9,18),
    ("lawn maintenance services huntsville al",11,18),("lawn aeration bartlett tn",10,17),
    ("lawn care memphis",3,17),("lawn fertilizer service bartlett tn",14,17),
    ("weed control services huntsville al",9,17),("lawn care olive branch ms",15,15),
    ("lawn care bartlett tn",13,15)]

if __name__ == "__main__":
    build()
