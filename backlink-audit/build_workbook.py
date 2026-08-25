#!/usr/bin/env python3
"""Consolidate the audit into a single-sheet Excel workbook.

Every evaluated item — domain-level verdicts plus the per-URL drill-down for
flagged domains — lands on one sheet so the analyst works from one surface.
The summary block uses live COUNTIFS/SUMIFS so counts update as
REVIEW_MANUALLY rows are resolved to DISAVOW or KEEP in place.

Usage: python3 build_workbook.py <outdir>
"""
import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from review_decisions import DECISIONS as REVIEW_RECOMMENDATIONS
except ImportError:          # the sheet still builds without the hand pass
    REVIEW_RECOMMENDATIONS = {}
from openpyxl.worksheet.datavalidation import DataValidation

# Columns in reading order: what it is, what to do, why, then the evidence,
# then volume, hosting and dates. Decision fields sit left so they stay
# visible while scrolling the evidence; the long Rationale goes last.
COL_GROUPS = [
    ("Identity", "Level", 9),
    ("Identity", "Referring Domain / URL", 42),
    ("Decision", "Action Recommendation", 23),
    ("Decision", "Remediation Priority", 30),
    ("Decision", "Confidence Score", 11),
    ("Decision", "Disavow Entry", 30),
    ("Why", "Primary Risk Factor", 42),
    ("Why", "Evidence Level", 30),
    ("Volume", "Backlinks (true)", 12),
    ("Volume", "Backlinks (in sample)", 12),
    ("Volume", "Follow (Equity) Links", 12),
    ("Volume", "Domain ascore", 11),
    ("Link evidence", "Anchor Text", 38),
    ("Link evidence", "Anchor Type", 14),
    ("Link evidence", "Target URL", 40),
    ("Link evidence", "Unique Anchors", 11),
    ("Link evidence", "Exact-Match Anchor %", 12),
    ("Link evidence", "Branded/URL Anchor %", 12),
    ("Link evidence", "Nofollow %", 10),
    ("Link evidence", "Avg External Links", 11),
    ("Link evidence", "Sponsored", 10),
    ("Link evidence", "Sitewide", 9),
    ("Link evidence", "Lost Link", 9),
    ("Link evidence", "Topically Relevant", 11),
    ("Hosting", "IP Address", 15),
    ("Hosting", "C-Block", 16),
    ("Hosting", "Hosting", 24),
    ("Hosting", "Country", 8),
    ("Dates", "First seen", 11),
    ("Dates", "Last seen", 11),
    ("Rationale", "Rationale", 95),
]
COLS = [(n, w) for _, n, w in COL_GROUPS]
GROUP_FILL = {"Identity": "1F3864", "Decision": "2E5A2E", "Why": "7B3F00",
              "Volume": "1F4E5F", "Link evidence": "4A3B6B",
              "Hosting": "5A5A5A", "Dates": "5A5A5A", "Rationale": "3F3F3F"}

FONT = "Arial"
ACTION_FILL = {
    "DISAVOW":               PatternFill("solid", fgColor="F8CBAD"),
    "REVIEW_MANUALLY":       PatternFill("solid", fgColor="FFE699"),
    "KEEP_AFFILIATE_RETAIN": PatternFill("solid", fgColor="C6E0B4"),
}
NUMERIC = {"Backlinks (true)", "Backlinks (in sample)",
           "Follow (Equity) Links", "Domain ascore", "Unique Anchors",
           "Avg External Links"}


TAB_COLS = [
    ("Referring Domain", 44), ("Disavow Entry", 30),
    ("Primary Risk Factor", 44), ("Confidence Score", 11),
    ("Evidence Level", 32), ("Remediation Priority", 32),
    ("Backlinks (true)", 13), ("Follow (Equity) Links", 13),
    ("Domain ascore", 11), ("C-Block", 17), ("Hosting", 24),
    ("Anchor Text", 36), ("First seen", 12), ("Last seen", 12),
    ("Rationale", 100),
]


PRIORITY_ORDER = {"P1": 0, "P2": 1, "P3": 2, "-": 3}


def _prio(row):
    return PRIORITY_ORDER.get(str(row.get("Remediation Priority", "-"))[:2], 3)


# Working-tab column order. The decision and the evidence behind it come
# first, then the numbers that would change your mind, then reference. An
# earlier build put Recommendation in column 16, which meant scrolling past
# fifteen columns to read the thing you are being asked to confirm.
REVIEW_COLS = [
    ("Referring Domain", 34), ("Recommendation", 15),
    ("Evidence for the recommendation", 74),
    ("Backlinks (true)", 10), ("Follow (Equity) Links", 9),
    ("Domain ascore", 8), ("Nofollow %", 9),
    ("Primary Risk Factor", 40), ("Anchor Text", 34),
    ("Disavow Entry", 30), ("Evidence Level", 30),
    ("First seen", 11), ("Last seen", 11), ("Rationale", 90),
]

DISAVOW_COLS = [
    ("Network", 42), ("Referring Domain", 34), ("Disavow Entry", 34),
    ("Remediation Priority", 30),
    ("Backlinks (true)", 10), ("Follow (Equity) Links", 9),
    ("Domain ascore", 8), ("Nofollow %", 9), ("Confidence Score", 10),
    ("Evidence Level", 32), ("Anchor Text", 32), ("C-Block", 16),
    ("Country", 8), ("First seen", 11), ("Last seen", 11),
    ("Rationale", 90),
]

REC_FILL = {"DISAVOW": "F8CBAD", "KEEP": "C6E0B4", "ASK_CLIENT": "FFE699"}
REC_ORDER = {"ASK_CLIENT": 0, "DISAVOW": 1, "KEEP": 2, "": 3}


def _work_tab(wb, title, rows, cols, blurb, confirm_header, confirm_list,
              default="", recommend=None, group_col=None):
    """A tab built to be worked top to bottom, not browsed."""
    ws = wb.create_sheet(title)
    HDR = 6
    first, last = HDR + 1, HDR + len(rows)
    n = len(cols)
    dec_col = get_column_letter(n + 1)
    ncol = n + 2

    ws["A1"] = f"{title} — {len(rows):,} domains"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="DDEBF7")
    ws["A2"] = blurb
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    ws["A3"] = "Backlinks covered"
    ws["B3"] = f"=SUM(${_col(cols,'Backlinks (true)')}${first}:"\
               f"${_col(cols,'Backlinks (true)')}${last})"
    ws["A4"] = "Rows still to decide"
    ws["B4"] = (f'=COUNTIF(${dec_col}${first}:${dec_col}${last},"{default}")'
                if default
                else f'=COUNTBLANK(${dec_col}${first}:${dec_col}${last})')
    for r in (3, 4):
        ws[f"A{r}"].font = Font(name=FONT, size=10, bold=True)
        ws[f"B{r}"].font = Font(name=FONT, size=10)
        ws[f"B{r}"].number_format = "#,##0"

    thin = Side(style="thin", color="BFBFBF")
    BOT, A_TOP = Border(bottom=thin), Alignment(vertical="top")
    A_WRAP = Alignment(vertical="top", wrap_text=True)
    F_BODY = Font(name=FONT, size=10)
    WRAPPED = {"Rationale", "Evidence for the recommendation", "Anchor Text"}

    heads = [c[0] for c in cols] + [confirm_header, "Notes"]
    widths = [c[1] for c in cols] + [24, 40]
    for i, (h, w) in enumerate(zip(heads, widths), start=1):
        c = ws.cell(HDR, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR].height = 30

    NUM = {"Backlinks (true)", "Follow (Equity) Links", "Domain ascore"}
    prev_group = None
    for ri, row in enumerate(rows, start=first):
        for ci, (name, _) in enumerate(cols, start=1):
            if name == "Recommendation":
                v = (recommend or {}).get(row["Referring Domain"],
                                          ("", ""))[0]
            elif name == "Evidence for the recommendation":
                v = (recommend or {}).get(row["Referring Domain"],
                                          ("", ""))[1]
            elif name == "Network":
                v = row.get("Primary Risk Factor", "")
            else:
                v = row.get(name, "")
            if name in NUM and v not in ("", None):
                try:
                    v = int(v)
                except ValueError:
                    pass
            c = ws.cell(ri, ci, v)
            c.font = F_BODY
            c.alignment = A_WRAP if name in WRAPPED else A_TOP
            c.border = BOT
            if name == "Recommendation" and v:
                c.font = Font(name=FONT, size=10, bold=True)
                c.fill = PatternFill("solid", fgColor=REC_FILL.get(v, "FFFFFF"))
            # A rule above each new network so 139 sibling domains read as
            # one block rather than 139 unrelated rows.
            if group_col and name == group_col:
                g = row.get("Primary Risk Factor", "")
                if g != prev_group:
                    c.border = Border(bottom=thin,
                                      top=Side(style="medium", color="1F3864"))
                    c.font = Font(name=FONT, size=10, bold=True)
        if group_col:
            prev_group = row.get("Primary Risk Factor", "")
        d = ws.cell(ri, n + 1, default)
        d.font = F_BODY
        d.fill = PatternFill("solid", fgColor="FFF2CC")
        d.border = BOT
        ws.cell(ri, n + 2).border = BOT

    ws.auto_filter.ref = f"A{HDR}:{get_column_letter(ncol)}{last}"
    ws.freeze_panes = ws.cell(first, 4).coordinate
    dv = DataValidation(type="list", formula1=confirm_list, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{dec_col}{first}:{dec_col}{last}")
    return ws


def _col(cols, name):
    return get_column_letter([c[0] for c in cols].index(name) + 1)


def _networks_tab(wb, dis):
    """The disavow list as the ~25 decisions it actually is.

    1,120 rows is not 1,120 judgements. 139 of them are seo-anomaly-s1.xyz
    through s139.xyz -- one operator, one call. This tab is the work surface;
    the Disavow tab is where you check a network's members before approving.
    """
    from collections import defaultdict
    g = defaultdict(list)
    for r in dis:
        g[r["Primary Risk Factor"]].append(r)
    groups = sorted(g.items(),
                    key=lambda kv: (min(_prio(x) for x in kv[1]),
                                    -sum(int(x["Follow (Equity) Links"] or 0)
                                         for x in kv[1]),
                                    kv[0]))

    ws = wb.create_sheet("2 Approve networks", 1)
    ws["A1"] = f"Approve by network — {len(groups)} decisions covering " \
               f"{len(dis):,} domains"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="DDEBF7")
    ws["A2"] = ("Highest-impact first: equity-passing networks at the top, "
                "nofollow-only at the bottom. Approve here, then spot-check "
                "the members on the Disavow tab — every domain in a row "
                "shares the signature named in Why it is flagged.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    HDR = 4
    heads = [("Priority", 26), ("Network", 44), ("Domains", 9),
             ("Backlinks", 11), ("Follow links", 13), ("Max authority", 12),
             ("Why it is flagged", 78), ("Example domain", 30),
             ("Approve? (Y / N / HOLD)", 22), ("Notes", 40)]
    for i, (h, w) in enumerate(heads, start=1):
        c = ws.cell(HDR, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR].height = 30

    thin = Side(style="thin", color="BFBFBF")
    BOT = Border(bottom=thin)
    F_BODY = Font(name=FONT, size=10)
    A_WRAP = Alignment(vertical="top", wrap_text=True)
    for ri, (name, members) in enumerate(groups, start=HDR + 1):
        members.sort(key=lambda r: -int(r["Backlinks (true)"]))
        prio = min(members, key=_prio)["Remediation Priority"]
        # A blank Follow column means the domain sits outside the link
        # sample, so its follow count is unknown -- not zero. Printing 0
        # there read as "passes no equity", which is the opposite of
        # "we cannot see".
        known = [m for m in members if m["Follow (Equity) Links"] != ""]
        fol = sum(int(m["Follow (Equity) Links"] or 0) for m in known)
        if not known:
            fol_cell = "not sampled"
        elif len(known) < len(members):
            fol_cell = f"{fol:,} (+{len(members) - len(known)} unsampled)"
        else:
            fol_cell = fol
        vals = [prio, name, len(members),
                sum(int(m["Backlinks (true)"]) for m in members),
                fol_cell,
                max(int(m["Domain ascore"] or 0) for m in members),
                members[0]["Rationale"], members[0]["Referring Domain"]]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(ri, ci, v)
            c.font = F_BODY
            c.alignment = A_WRAP if ci in (2, 7) else Alignment(vertical="top")
            c.border = BOT
            if ci in (3, 4, 6) or (ci == 5 and isinstance(v, int)):
                c.number_format = "#,##0"
        ws.cell(ri, 2).font = Font(name=FONT, size=10, bold=True)
        d = ws.cell(ri, 9, "")
        d.fill = PatternFill("solid", fgColor="FFF2CC")
        d.border = BOT
        ws.cell(ri, 10).border = BOT

    last = HDR + len(groups)
    ws.cell(last + 2, 2, "TOTAL").font = Font(name=FONT, size=10, bold=True)
    for ci, col in ((3, "C"), (4, "D")):
        c = ws.cell(last + 2, ci, f"=SUM({col}{HDR+1}:{col}{last})")
        c.font = Font(name=FONT, size=10, bold=True)
        c.number_format = "#,##0"
    ws.freeze_panes = ws.cell(HDR + 1, 3).coordinate
    ws.auto_filter.ref = f"A{HDR}:J{last}"
    dv = DataValidation(type="list", formula1='"Y,N,HOLD"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I{HDR+1}:I{last}")
    return len(groups)


def _decisions_tab(wb, dom):
    """Audit trail for the queue once the client has cleared it.

    Worth its own tab because on 37 of these the client's call and the
    audit's recommendation differ, and that disagreement is a fact about the
    file someone will want to see later -- not something to quietly overwrite
    so the two agree.
    """
    try:
        from review_decisions import DECISIONS
    except ImportError:
        DECISIONS = {}
    rows = [r for r in dom
            if "Client Decision" in r["Primary Risk Factor"]
            or "Client Decision" in r["Rationale"][:40]]
    rows = [r for r in dom if r["Referring Domain"] in DECISIONS]
    order = {"DISAVOW": 0, "KEEP_AFFILIATE_RETAIN": 1}
    rows.sort(key=lambda r: (order.get(r["Action Recommendation"], 2),
                             -int(r["Follow (Equity) Links"] or 0),
                             r["Referring Domain"]))
    ws = wb.create_sheet("Decisions log", 3)
    ws["A1"] = ("REFERENCE — decisions log: "
                f"{len(rows)} domains from the review queue")
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="EDEDED")
    ws["A2"] = ("The review queue, resolved. 'Audit said' is what this audit "
                "recommended and 'Applied' is the decision in the file, so "
                "where the two differ it stays on the record. Reversing any "
                "row means deleting one line from disavow_full.txt.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    HDR = 4
    heads = [("Referring Domain", 34), ("Applied", 24), ("Audit said", 13),
             ("Agree?", 9), ("Backlinks", 10), ("Follow links", 11),
             ("Authority", 10), ("Nofollow %", 10),
             ("Why the audit said that", 84), ("Notes", 36)]
    for i, (h, w) in enumerate(heads, start=1):
        c = ws.cell(HDR, i, h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR].height = 30

    thin = Side(style="thin", color="BFBFBF")
    BOT = Border(bottom=thin)
    F_BODY = Font(name=FONT, size=10)
    A_WRAP = Alignment(vertical="top", wrap_text=True)
    for ri, r in enumerate(rows, start=HDR + 1):
        mine = DECISIONS.get(r["Referring Domain"], ("", ""))[0]
        why = DECISIONS.get(r["Referring Domain"], ("", ""))[1]
        applied = ("DISAVOW" if r["Action Recommendation"] == "DISAVOW"
                   else "KEEP")
        agree = "" if not mine else ("yes" if mine == applied else "NO")
        vals = [r["Referring Domain"], r["Primary Risk Factor"], mine, agree,
                int(r["Backlinks (true)"]),
                int(r["Follow (Equity) Links"] or 0)
                if r["Follow (Equity) Links"] != "" else "not sampled",
                int(r["Domain ascore"] or 0), r["Nofollow %"], why]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(ri, ci, v)
            c.font = F_BODY
            c.alignment = A_WRAP if ci in (2, 9) else Alignment(vertical="top")
            c.border = BOT
            if ci in (5, 6, 7) and isinstance(v, int):
                c.number_format = "#,##0"
        ws.cell(ri, 2).fill = PatternFill(
            "solid", fgColor="F8CBAD" if applied == "DISAVOW" else "C6E0B4")
        a = ws.cell(ri, 4)
        if agree == "NO":
            a.font = Font(name=FONT, size=10, bold=True, color="9C0006")
            a.fill = PatternFill("solid", fgColor="FFC7CE")
        ws.cell(ri, 10).border = BOT
    last = HDR + len(rows)
    ws.freeze_panes = ws.cell(HDR + 1, 2).coordinate
    ws.auto_filter.ref = f"A{HDR}:J{last}"
    return ws


def _start_tab(wb, dom, n_networks):
    from collections import Counter
    c = Counter(r["Action Recommendation"] for r in dom)
    dis = [r for r in dom if r["Action Recommendation"] == "DISAVOW"]
    follow = sum(int(r["Follow (Equity) Links"] or 0) for r in dis)
    nf_only = sum(1 for r in dis
                  if r["Remediation Priority"].startswith("P3"))
    ws = wb.create_sheet("1 Start here", 0)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 104
    row = [2]

    def w(b="", cc="", bold=False, size=10, colour="000000", gap=0):
        r = row[0]
        if b:
            ws.cell(r, 2, b).font = Font(name=FONT, size=size, bold=bold,
                                         color=colour)
        if cc:
            cl = ws.cell(r, 3, cc)
            cl.font = Font(name=FONT, size=size, color=colour)
            cl.alignment = Alignment(vertical="top", wrap_text=True)
        row[0] = r + 1 + gap

    w("Performance Lab — backlink audit", size=16, bold=True, gap=1)
    w("What this is",
      f"{len(dom):,} referring domains, every one carrying a verdict. "
      "1,362,105 backlinks, reconciled to Semrush.", bold=True, gap=1)

    w("Where things stand", "", bold=True)
    w("Disavow", f"{c['DISAVOW']:,} domains — {follow:,} follow links between "
      f"them. {nf_only:,} are nofollow-only, so they pass no equity.")
    if c["REVIEW_MANUALLY"]:
        w("Needs your call",
          f"{c['REVIEW_MANUALLY']} domains. Each one carries my "
          "recommendation and the evidence for it.")
    else:
        w("Needs your call",
          "Nothing outstanding — the review queue is cleared. The "
          "Decisions log tab records what was decided and where it differs "
          "from what this audit recommended.")
    w("Keep", f"{c['KEEP_AFFILIATE_RETAIN']:,} domains — affiliates, brand "
      "estate, editorial citations and low-exposure retains.", gap=1)

    w("The tabs", "", bold=True)
    w("How to read them",
      "Numbered tabs are the workflow, in order. The two unnumbered tabs are "
      "reference — nothing in them needs a decision.")
    w("2 Approve networks",
      f"Start here. The {c['DISAVOW']:,} disavows are {n_networks} networks, "
      f"not {c['DISAVOW']:,} separate judgements — 139 rows are "
      "seo-anomaly-s1.xyz through s139.xyz, one operator. Approve or reject "
      "each network. "
      "Equity-passing ones are at the top; nofollow-only at the bottom, and "
      "those change nothing either way.")
    w("3 Disavow list",
      "The members of each network, grouped under it with a rule between "
      "groups. Use it to spot-check a network before approving, or to pull a "
      "single domain out of one.")
    if c["REVIEW_MANUALLY"]:
        w("3b Review queue",
          f"The {c['REVIEW_MANUALLY']} the rules would not decide. "
          "Recommendation and evidence sit in columns B and C. Set the "
          "Decision column.")
    else:
        w("Decisions log  (reference)",
          "The review queue after your call: 7 retained by name, the other "
          "63 disavowed as a block. 'Audit said' keeps this audit's own "
          "recommendation next to what was applied, and the Agree column "
          "flags the 38 rows where the two differ. Reversing any of them is "
          "one line out of disavow_full.txt.")
    w("All domains  (reference)",
      f"Every one of the {len(dom):,} domains, with each sampled backlink "
      "listed under its domain. Filter Level=DOMAIN for the verdict list. "
      "Nothing to action.", gap=1)

    dtcx = next((r for r in dom if r["Referring Domain"] == "dtcx.com"), None)
    decided = bool(dtcx) and dtcx["Action Recommendation"] == "DISAVOW"
    w("Still worth confirming" if decided else "One question for you",
      "", bold=True, colour="9C3A00")
    w("dtcx.com",
      "It links to you with image anchors \"Performance Lab Logo\" and "
      "\"Nutropic Logo\", which reads like your own or a partner's property. "
      "But it is also promoted BY several of the spam networks in this audit "
      "— \"visit dtcx.com for latest info\" on two of them, and \"Premium PBN "
      "Network Service dtcx.com Rank First\" on a link vendor's page. Either "
      "it is yours, or a seller is riding the brand."
      + (" It is now disavowed with the rest of the queue; if it turns out to "
         "be yours, delete \"domain:dtcx.com\" from the file before you "
         "submit." if decided else
         " I have left it unresolved rather than guess."),
      colour="9C3A00", gap=1)

    w("What to submit", "", bold=True)
    w("The file",
      "disavow_full.txt is the file for Google's tool — every disavowed "
      "domain, grouped by network with comment headers. disavow_core.txt is "
      "the equity-passing subset and disavow_nofollow_hygiene.txt the inert "
      "remainder; together they are exactly disavow_full.txt. Submitting the "
      "full file matches your call to include nofollow.", gap=1)

    w("Where the confidence sits", "", bold=True)
    ev = Counter(r["Evidence Level"] for r in dom)
    obs = sum(v for k, v in ev.items() if k.startswith("Link-level"))
    w("Observed placements", f"{obs:,} domains judged on their actual links — "
      "anchors, targets, page URLs.")
    w("Domain metrics only",
      f"{ev['Domain-level only (outside sample)']:,} domains fall outside the "
      "link sample, so authority, hosting, name and volume are all there was.")
    tri = sum(1 for r in dom if "Negligible Exposure" in r["Primary Risk Factor"])
    w("Retained on exposure",
      f"{tri:,} keeps are a risk decision, not a clean bill of health: too "
      "few links to be worth either a disavow or your time. They are "
      "labelled as such in the Full audit tab.")
    w("Known gaps",
      "appsrankings.com (22,204 links) is retained as affiliate on inference, "
      "not an observed redirect — one filtered Semrush pull would settle it. "
      "Topical relevance is scored on an English lexicon, so non-English "
      "sources are under-detected. SUMMARY.md lists a doorway-URL pattern "
      "found but deliberately not automated.")
    return ws


def main(outdir):
    dom = list(csv.DictReader(open(f"{outdir}/full_refdomain_audit.csv", encoding="utf-8")))
    url = list(csv.DictReader(open(f"{outdir}/url_drilldown.csv", encoding="utf-8")))
    order = {"DISAVOW": 0, "REVIEW_MANUALLY": 1, "KEEP_AFFILIATE_RETAIN": 2}
    dom.sort(key=lambda r: (order[r["Action Recommendation"]],
                            -int(r["Backlinks (true)"])))

    by_domain = {}
    for r in url:
        by_domain.setdefault(r["Referring Domain"], []).append(r)

    rows = []
    for d in dom:
        key = d["Referring Domain"]
        rows.append({**d, "Level": "DOMAIN",
                     "Referring Domain / URL": key,
                     "Anchor Type": "", "Sponsored": "",
                     "Sitewide": "", "Lost Link": ""})
        for u in by_domain.get(key, []):
            rows.append({
                "Level": "URL",
                "Referring Domain / URL": u["Referring Domain / URL"],
                "Target URL": u["Target URL"], "Anchor Text": u["Anchor Text"],
                "Anchor Type": u["Anchor Type"],
                # A URL inherits its domain's final verdict. It used to
                # carry the link-level classifier's own reading, which is an
                # earlier stage: 785 URL rows still said REVIEW_MANUALLY
                # after the domain pass and the client had resolved every
                # domain. One verdict per domain, and the domain pass owns it.
                "Action Recommendation": d["Action Recommendation"],
                "Primary Risk Factor": u["Primary Risk Factor"],
                "Confidence Score": u["Confidence Score"],
                "Domain ascore": u["Page AScore"],
                "Avg External Links": u["External Links"],
                "Nofollow %": "100%" if u["Nofollow"] == "true" else "0%",
                "Sponsored": u["Sponsored"], "Sitewide": u["Sitewide"],
                "Lost Link": u["Lost Link"], "First seen": u["First Seen"],
                "Last seen": u["Last Seen"],
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "All domains"
    HDR = 9
    first, last = HDR + 1, HDR + len(rows)
    act, lvl = f"$C${first}:$C${last}", f"$A${first}:$A${last}"
    bl, fol = f"$I${first}:$I${last}", f"$K${first}:$K${last}"

    ws["A1"] = ("REFERENCE — every domain and every sampled URL. "
                "Nothing to action here; use it to look something up.")
    ws["A1"].font = Font(name=FONT, size=15, bold=True)
    ws["A2"] = (
        "One row per referring domain (Level=DOMAIN), plus a row per sampled "
        "backlink underneath it (Level=URL). A URL row shows the same verdict "
        "as its domain — the domain is the unit of decision, so filtering "
        "Level=DOMAIN gives you the verdict list with no double counting. "
        "Evidence Level says whether the call rests on observed placements or "
        "on domain metrics alone.")
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

    for col, head in zip("ABCD", ["Action", "Domains", "Backlinks (true)",
                                  "Follow links (sampled)"]):
        ws[f"{col}4"] = head
    # Only the verdicts actually present. Once the review queue is cleared a
    # REVIEW_MANUALLY row is a live formula reading zero, which is the one
    # thing on the sheet that looks like outstanding work when there is none.
    present = [a for a in ("DISAVOW", "REVIEW_MANUALLY",
                           "KEEP_AFFILIATE_RETAIN")
               if any(x["Action Recommendation"] == a for x in dom)]
    for i, a in enumerate(present):
        r = 5 + i
        ws[f"A{r}"] = a
        ws[f"B{r}"] = f'=COUNTIFS({act},$A{r},{lvl},"DOMAIN")'
        ws[f"C{r}"] = f'=SUMIFS({bl},{act},$A{r},{lvl},"DOMAIN")'
        ws[f"D{r}"] = f'=SUMIFS({fol},{act},$A{r},{lvl},"DOMAIN")'
        ws[f"A{r}"].fill = ACTION_FILL[a]
    tot = 5 + len(present)
    ws[f"A{tot}"] = "TOTAL"
    for col in "BCD":
        ws[f"{col}{tot}"] = f"=SUM({col}5:{col}{tot - 1})"
    for r in range(4, tot + 1):
        for c in "ABCD":
            cell = ws[f"{c}{r}"]
            cell.font = Font(name=FONT, size=10, bold=(r in (4, 8)))
            if c != "A":
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    thin = Side(style="thin", color="BFBFBF")
    BOT = Border(bottom=thin)
    F_DOM = Font(name=FONT, size=10, bold=True, color="000000")
    F_URL = Font(name=FONT, size=10, bold=False, color="404040")
    A_TOP = Alignment(vertical="top")
    A_WRAP = Alignment(vertical="top", wrap_text=True)
    WRAPPED = {"Rationale", "Primary Risk Factor"}
    DOM_FILL = PatternFill("solid", fgColor="D9E1F2")
    # Group band one row above the names, merged per group, so 31 columns
    # read as six sections rather than an undifferentiated wall.
    _gi = 1
    while _gi <= len(COL_GROUPS):
        _grp = COL_GROUPS[_gi - 1][0]
        _span = 1
        while (_gi + _span <= len(COL_GROUPS)
               and COL_GROUPS[_gi + _span - 1][0] == _grp):
            _span += 1
        if _span > 1:
            ws.merge_cells(start_row=HDR - 1, start_column=_gi,
                           end_row=HDR - 1, end_column=_gi + _span - 1)
        _gc = ws.cell(row=HDR - 1, column=_gi, value=_grp.upper())
        _gc.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        _gc.fill = PatternFill("solid", fgColor=GROUP_FILL[_grp])
        _gc.alignment = Alignment(horizontal="center", vertical="center")
        _gi += _span

    for i, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=HDR, column=i, value=name)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GROUP_FILL[COL_GROUPS[i - 1][0]])
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[HDR - 1].height = 16
    ws.row_dimensions[HDR].height = 30

    for ri, row in enumerate(rows, start=first):
        is_dom = row["Level"] == "DOMAIN"
        for ci, (name, _) in enumerate(COLS, start=1):
            v = row.get(name, "")
            if name in NUMERIC and v not in ("", None):
                try:
                    v = float(v) if "." in str(v) else int(v)
                except ValueError:
                    pass
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = F_DOM if is_dom else F_URL
            c.alignment = A_WRAP if name in WRAPPED else A_TOP
            c.border = BOT
        ws.cell(row=ri, column=3).fill = ACTION_FILL[row["Action Recommendation"]]
        if is_dom:
            ws.cell(row=ri, column=1).fill = DOM_FILL

    ws.auto_filter.ref = f"A{HDR}:{get_column_letter(len(COLS))}{last}"
    ws.freeze_panes = f"C{first}"
    dv = DataValidation(type="list",
                        formula1='"DISAVOW,REVIEW_MANUALLY,KEEP_AFFILIATE_RETAIN"',
                        allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"C{first}:C{last}")

    # ---- Disavow: grouped by network, equity first ----------------------
    dis = [dict(d) for d in dom if d["Action Recommendation"] == "DISAVOW"]
    for r in dis:
        r["Disavow Entry"] = r.get("Disavow Entry") or \
            f"domain:{r['Referring Domain']}"
    # Order networks exactly as the Networks tab does, then keep each one
    # contiguous. Sorting by priority first split "Numbered Sibling Domain
    # Network" across two bands and "Hacked Site" across three, so a network
    # approved on the Networks tab had members scattered down the sheet.
    _net_rank = {}
    for r in dis:
        k = r["Primary Risk Factor"]
        _net_rank[k] = min(_net_rank.get(k, 99), _prio(r))
    _net_follow = {}
    for r in dis:
        k = r["Primary Risk Factor"]
        _net_follow[k] = _net_follow.get(k, 0) + int(
            r["Follow (Equity) Links"] or 0)
    dis.sort(key=lambda r: (_net_rank[r["Primary Risk Factor"]],
                            -_net_follow[r["Primary Risk Factor"]],
                            r["Primary Risk Factor"],
                            _prio(r), -int(r["Backlinks (true)"])))
    _work_tab(
        wb, "3 Disavow list", dis, DISAVOW_COLS,
        "Grouped by network, equity-passing first. 'Disavow Entry' is the "
        "exact line for Google's tool — note where it names a subdomain: "
        "those hosts are compromised, not hostile, and the narrow scope is "
        "deliberate. Approve whole networks on the Networks tab; use this to "
        "check members or pull one out.",
        confirm_header="Confirmed? (Y / N / HOLD)",
        confirm_list='"Y,N,HOLD"',
        group_col="Network",
    )

    # ---- Review queue: recommendation and evidence up front -------------
    rev = [dict(d) for d in dom
           if d["Action Recommendation"] == "REVIEW_MANUALLY"]
    for r in rev:
        r["Disavow Entry"] = f"domain:{r['Referring Domain']}"
    rev.sort(key=lambda r: (
        REC_ORDER.get(REVIEW_RECOMMENDATIONS.get(
            r["Referring Domain"], ("", ""))[0], 3),
        -int(r["Backlinks (true)"])))
    if not rev:
        _decisions_tab(wb, dom)
    else:
        _work_tab(
            wb, "3b Review queue", rev, REVIEW_COLS,
        "What the rules would not decide, read one by one against the actual "
        "placements. My recommendation is column B and the evidence for it "
        "column C. Ordered: the one open question first, then disavows, then "
        "keeps. Set the Decision column — disagreeing with me is the point.",
            confirm_header="Decision (DISAVOW / KEEP / PENDING)",
            confirm_list='"DISAVOW,KEEP,PENDING"',
            default="PENDING",
            recommend=REVIEW_RECOMMENDATIONS,
        )

    n_net = _networks_tab(wb, [dict(d) for d in dis])
    _start_tab(wb, dom, n_net)
    # Reading order: brief, then the three worklists, then the reference set.
    wb.move_sheet("All domains", offset=len(wb.sheetnames))
    wb.active = 0

    path = f"{outdir}/performancelab_backlink_audit.xlsx"
    wb.save(path)
    print(f"{len(rows):,} rows -> {path}")
    print(f"  DOMAIN rows: {sum(1 for r in rows if r['Level']=='DOMAIN'):,}")
    print(f"  URL rows   : {sum(1 for r in rows if r['Level']=='URL'):,}")


if __name__ == "__main__":
    main(sys.argv[1])
