#!/usr/bin/env python3
"""Re-derive classification + independent QA verification + multi-tab workbook.
QA pass critically checks EACH domain: does the decision hold up, and which
decisions are borderline enough to need a human second look."""
import csv, re, datetime, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def ts(v):
    try: return datetime.datetime.utcfromtimestamp(int(v)).strftime('%Y-%m-%d')
    except: return ''
def isots(v): return (v or '')[:10]

# ---------- load raw ----------
ah={}
for r in csv.DictReader(open('ahrefs_refdomains.csv')):
    d=r['domain'].strip().lower()
    if not d: continue
    ah[d]=dict(dr=float(r['domain_rating'] or 0),traffic=int(r['traffic_domain'] or 0),
        links=int(r['links_to_target'] or 0),dofollow=int(r['dofollow_links'] or 0),
        is_spam=str(r['is_spam']).lower()=='true',positions=int(r['positions_source_domain'] or 0),
        ip=(r['ip_source'] or '').strip(),first=isots(r['first_seen']),last=isots(r['last_seen']))
sm={}
for r in csv.DictReader(open('semrush_refdomains.csv'),delimiter=';'):
    d=r['domain'].strip().lower()
    if not d: continue
    sm[d]=dict(ascore=int(r['domain_score'] or 0),trust=int(r['domain_trust_score'] or 0),
        links=int(r['backlinks_num'] or 0),ip=(r['ip'] or '').strip(),country=(r['country'] or '').strip(),
        first=ts(r['first_seen']),last=ts(r['last_seen']))
all_domains=sorted(set(ah)|set(sm))

# ---------- link-network IPs (CDN/cloud excluded) ----------
CDN_PREFIXES=('104.16.','104.17.','104.18.','104.19.','104.20.','104.21.','104.22.','104.23.',
 '104.24.','104.25.','104.26.','104.27.','104.28.','172.64.','172.65.','172.66.','172.67.','172.68.',
 '172.69.','172.70.','172.71.','188.114.','162.159.','131.0.72.','151.101.','142.250.','142.251.',
 '172.253.','173.194.','64.233.','192.178.','192.179.','108.177.','150.171.','66.102.','216.58.',
 '34.','35.','3.','13.','18.','20.','52.','54.','23.','2.','5.161.')
def is_cdn(ip): return ip.startswith(CDN_PREFIXES)
ipmap=collections.defaultdict(list)
for d in all_domains:
    ip=(sm.get(d,{}).get('ip') or ah.get(d,{}).get('ip') or '').strip()
    if ip and not is_cdn(ip): ipmap[ip].append(d)
network_ips={ip:doms for ip,doms in ipmap.items() if len(doms)>=4}

# ---------- vocab ----------
WHITELIST_BRAND={'apple.com','yahoo.com','bing.com','pinterest.com','bbb.org','yellowpages.com',
 'nextdoor.com','zoominfo.com','glass.com','glassonweb.com','chamberofcommerce.com','porch.com',
 'constantcontact.com','constantcontactpages.com','superpages.com','dexknows.com','expertise.com',
 'fixr.com','housedigest.com','moneytalksnews.com','owler.com','inkl.com','barbend.com','accio.com',
 'seamless.ai','castbox.fm','coroflot.com','brightside.me','diamondcertified.org','qualifiedremodeler.com',
 'windowanddoor.com','windowdigest.com','dwmmag.com','growjo.com','enigma.com','bloomberry.com',
 'neverbounce.com','arounddeal.com','wiza.co','conves.io','weblancer.net','re-thinkingthefuture.com',
 '24-7pressrelease.com','vsbattles.com','processregister.com','growthzoneapp.com','williamsonchamber.com',
 'infinitywindows.com','energyswingwindows.com','parse.gl','glarity.app','tntcode.com','hometownstation.com'}
# B2B data aggregators / tool sites that auto-list every company -> harmless links
AGGREGATORS={'zoominfo.com','owler.com','growjo.com','seamless.ai','arounddeal.com','wiza.co',
 'neverbounce.com','enigma.com','bloomberry.com','dexknows.com','superpages.com','yellowpages.com',
 'expertise.com','porch.com','fixr.com','chamberofcommerce.com','bbb.org','nextdoor.com','siteprice.org',
 'sitelike.org','csswinner.com','coroflot.com','processregister.com','conves.io','accio.com','castbox.fm',
 'glarity.app','parse.gl','bizhwy.com','smb.co','arounddeal.com'}
FREE_HOSTS=('.blogspot.com','.pages.dev','.azurewebsites.net','.freemyip.com','.web.app','.wordpress.com')
def m(d,p): return re.search(p,d) is not None
LINK_SELLER=r'(backlink|baclink|seolink|seobacklink|buyseolinks|linkbuy|buylink|guestpost|rankvance|qualitybacklinks|welinkbacklinks|linksnatcher|daolink|skylinkseo|trafficbooster|increasewebtraffic|worldbusinesspromote|goooogla|newblogerseo|blogerreviewers|mediaboooster|clicktobuy|quickseolinks|rarebacklinks|superbquality|viralbacklinks|eliteseo|bestrank|bestsites|buyrank|buyfair|booastranking|friendlybacklinks|webranking|topratedbacklinks|allbaclinks|atozbacklinks|99backlinks|blinks\.)'
GAMBLING=r'(casino|poker|betwinner|ufabet|slot|pick4d|truebookies|m98ufa|\bbet\b|4d\.live|onlinegaming|hotonline)'
ADULT=r'(porn|xxx|\bsex\b|ineed2pee|submissive|dominated)'
STATSPAM=r'(webstats|websiteworth|worthchecker|domainanalysis|siteprice|sitelike|domaineye|gsitestatus|seodomains|websitescrawl|websiterace|domainsc\.com|getwebsiteworth|webworthchecker|complaintinfo)'
DIR_SPAM=r'(webdirectory|allwebsitesdirectory|newwebsiteslist|alltopleveldomains|topleveldomains|yesdomains|all-aged-domains|domainwork|domains\.space|uklistingz|toplist\.co|01webdirectory|trendspotdirectory|directorylinkservice)'
SHORTENER=r'(shorten|shrink|byteshort|anchorurl|atomizelink|urls-shortener|screenshots\.wiki|buzzshrink|\.icu$)'
DOMRAIDER=r'(domraider|expireddomain|all-aged-domains|alltopleveldomains)'
MOLDOVA_IP='195.20.19.178'
FARM_IPS={'203.161.54.114','118.139.181.85','118.139.176.46','118.139.178.200','118.139.161.199',
 '118.139.181.255','184.168.115.60','184.168.111.168','184.168.116.137','184.168.113.185',
 '118.139.178.110','118.139.177.45','184.168.108.151','184.168.109.253','184.168.107.83'}
TOPICAL=r'(window|door|glass|home|house|roof|contractor|remodel|renov|exterior|build|construct|shutter|siding|energy|hvac|atlanta|georgia|roswell|nashville|realestate|property|residence|kitchen|bath|decor|interior|garage|fence|deck|patio)'

# ---------- curated human judgment (boundary domains) ----------
CURATED={
 'disgustingmen.com':'DR39 men’s-interest magazine; a windows link here is off-topic — likely paid/guest placement. Verify naturalness.',
 'enthrallinggumption.com':'Generic spun-content blog name (DR35), off-topic — possible high-DR PBN. Verify before trusting.',
 'vsbattles.com':'Fan wiki (DR33); link is likely a forum/profile mention — low risk, keep.',
 'castbox.fm':'Podcast platform profile/embed — harmless.',
 'pinuphouses.com':'Tiny-house-plans site (DR34), topically adjacent — likely genuine.',
 'ahouseinthehills.com':'Real home/lifestyle blog — plausibly genuine editorial/sponsored.',
 'sometimes-homemade.com':'Real food/home blog (DR47) — likely genuine.',
 'lovehappensmag.com':'Lifestyle magazine (DR38) — likely genuine or sponsored.',
 'thefuntimesguide.com':'Long-running content site — likely genuine.',
 'roswell365.com':'Local Roswell GA site — genuine local citation.',
 'atlantahomeimprovement.com':'Local home-improvement magazine (91 links) — likely genuine; check links aren’t paid/sitewide.',
 'derchidoor.com':'Door-related, 143 links = probable sitewide/widget link — verify placement.',
 'fairviewwindows.co.uk':'Window company peer/competitor — genuine, keep.',
 'secureglaze-windows.co.uk':'Window company peer — genuine, keep.',
 'windorpro.co.za':'Window/door company — genuine, keep.',
 'fallatlantahomeshow.com':'Local home-show event site — genuine citation.',
 'nashvillefallhomeshow.com':'Local home-show event site — genuine citation.',
 'roswellartfestival.com':'Local Roswell GA event — genuine citation.',
 'longislandwindowsandsiding.com':'Window/siding company — genuine peer.',
 'siteprice.org':'Auto-generated “website value” aggregator — harmless, disavow optional.',
 'sitelike.org':'Auto-generated “similar sites” aggregator — harmless, disavow optional.',
 'csswinner.com':'Design-award directory — low value but not toxic; optional.',
}

rows=[]
for d in all_domains:
    a=ah.get(d,{}); s=sm.get(d,{})
    dr=a.get('dr',0.0); asc=s.get('ascore',0); auth=max(dr,asc)
    traffic=a.get('traffic',0); positions=a.get('positions',0); is_spam=a.get('is_spam',False)
    ip=(s.get('ip') or a.get('ip') or '').strip(); country=s.get('country','')
    links=max(a.get('links',0),s.get('links',0))
    src=[x for x in (('ahrefs' if d in ah else None),('semrush' if d in sm else None)) if x]
    sources='+'.join(src)
    topical='Y' if m(d,TOPICAL) else ''
    is_free=d.endswith(FREE_HOSTS)

    footprint=None
    if m(d,LINK_SELLER): footprint='Link-selling / SEO-service PBN'
    elif d.endswith('.blogspot.com'): footprint='Blogspot PBN / spam post'
    elif m(d,GAMBLING): footprint='Gambling / casino spam'
    elif m(d,ADULT): footprint='Adult spam'
    elif m(d,SHORTENER) or ip==MOLDOVA_IP: footprint='URL-shortener / redirect link network'
    elif m(d,DOMRAIDER): footprint='Expired-domain / auto network'
    elif m(d,STATSPAM): footprint='Auto-generated stats / worth-checker'
    elif m(d,DIR_SPAM): footprint='Low-quality directory / TLD list'
    elif is_free: footprint='Free-host throwaway page'
    hard_toxic=footprint in ('Link-selling / SEO-service PBN','Gambling / casino spam','Adult spam',
        'URL-shortener / redirect link network','Blogspot PBN / spam post','Expired-domain / auto network')

    reasons=[]; signals=[]; cat=''; conf=''
    if d in WHITELIST_BRAND or (auth>=30 and not is_free and not hard_toxic and ip!=MOLDOVA_IP):
        cat='LEGITIMATE / high-authority'; conf='KEEP'
        reasons.append(f'authority={auth:.0f} (DR{dr:.0f}/AS{asc}); recognized or strong domain')
    else:
        if is_spam: signals.append('Ahrefs is_spam=true')
        if ip in network_ips and ip: signals.append(f'link-network IP {ip} (+{len(network_ips[ip])-1} siblings)')
        if ip in FARM_IPS: signals.append('known link-farm IP block')
        if country=='md': signals.append('Moldova spam cluster')
        if auth<=3 and traffic==0 and positions==0: signals.append(f'dead (auth={auth:.0f},0 traffic,0 kw)')
        elif traffic==0 and positions==0 and auth<=6: signals.append(f'no organic traffic/kw (auth={auth:.0f})')
        nsig=len(signals)
        if footprint:
            cat=footprint
            if footprint in ('Auto-generated stats / worth-checker','Low-quality directory / TLD list','Free-host throwaway page'):
                conf='HIGH' if nsig>=1 else 'MEDIUM'
            else: conf='HIGH'
            reasons.append('footprint: '+footprint)
        elif nsig>=2: cat='Multi-signal spam'; conf='HIGH'; reasons.append('2+ independent toxic signals')
        elif nsig==1: cat='Single-signal low-quality'; conf='MEDIUM'; reasons.append('one strong toxic signal')
        elif auth<=6: cat='Low-authority (needs review)'; conf='REVIEW'; reasons.append(f'low authority={auth:.0f}')
        else: cat='Unclear'; conf='REVIEW'; reasons.append(f'authority={auth:.0f}; borderline')
        reasons.extend(signals)

    decision={'HIGH':'DISAVOW','MEDIUM':'DISAVOW','REVIEW':'REVIEW','KEEP':'KEEP'}[conf]

    # ---- window-company microsite cluster (shared AWS IPs; near-identical brand variants) ----
    MICROSITE_IPS={'15.197.225.128','15.197.142.173','3.33.152.147','3.33.251.168'}
    if (ip in MICROSITE_IPS and m(d,r'(window|door)')) or d=='windowdoor-test.com':
        cat='Window-company microsite cluster (verify OWNERSHIP)'
        conf='REVIEW'; decision='REVIEW'
        reasons=['near-identical window-brand domain on a shared AWS IP with other window microsites',
                 f'auth={auth:.0f}, dead — likely the client’s own microsite/typo-domain or a self-built PBN']

    # ---------- INDEPENDENT QA / second-look pass ----------
    qa=[]
    if sources=='ahrefs+semrush' and conf in ('HIGH','MEDIUM'):
        qa.append('CONFIRMED by both Ahrefs+Semrush')
    if conf=='KEEP' and d not in WHITELIST_BRAND and d not in AGGREGATORS:
        if not topical: qa.append('KEEP on authority only + off-topic — verify editorial vs paid guest post')
        else: qa.append('KEEP on authority — topically relevant, likely genuine')
    if conf in ('KEEP','REVIEW') and (is_spam or ip in FARM_IPS or ip in network_ips or country=='md'):
        qa.append('PASSED but carries a toxic signal — reconsider for disavow')
    if conf=='MEDIUM' and signals and all('traffic' in x or 'dead' in x for x in signals) and (topical or m(d,r'(inc|llc|co\.|company|group|pro|services)')) and links>=3:
        qa.append('MEDIUM on “dead” signal only + business-like name — verify it is not a real small/local site')
    if conf in ('HIGH','MEDIUM') and (traffic>200 or positions>50):
        qa.append(f'flagged but has real organic footprint (traffic={traffic}, kw={positions}) — double-check')
    if conf=='REVIEW' and (m(d,LINK_SELLER) or m(d,GAMBLING) or m(d,STATSPAM) or m(d,DIR_SPAM)):
        qa.append('REVIEW but name looks spammy — consider disavow')
    if 'microsite cluster' in cat:
        qa.append('OWNERSHIP CHECK: confirm whether you own this. Do NOT disavow your own sites — redirect/consolidate instead. Disavow only if it is a third-party PBN.')
    if d in CURATED: qa.append('MANUAL: '+CURATED[d])

    rows.append(dict(domain=d,sources=sources,decision=decision,confidence=conf,category=cat,
        dr=round(dr,1),ascore=asc,traffic=traffic,positions=positions,is_spam='TRUE' if is_spam else '',
        links=links,ip=ip,country=country,first=s.get('first') or a.get('first',''),
        last=s.get('last') or a.get('last',''),topical=topical,signals=len(signals),
        evidence=' | '.join(reasons),qa_note=' ; '.join(qa)))

# ---------- regenerate disavow.txt ----------
flagged=[r for r in rows if r['confidence'] in ('HIGH','MEDIUM')]
high=[r for r in flagged if r['confidence']=='HIGH']; med=[r for r in flagged if r['confidence']=='MEDIUM']
with open('disavow.txt','w') as o:
    o.write(f'# Disavow file for ngwindows.com  ({datetime.date.today()})\n')
    o.write('# From Ahrefs + Semrush referring-domain audit. Review before submitting.\n')
    o.write(f'# HIGH: {len(high)}  MEDIUM: {len(med)}  (total {len(high)+len(med)})\n#\n')
    for label,grp in (('HIGH CONFIDENCE',high),('MEDIUM CONFIDENCE (verify)',med)):
        o.write(f'# ===== {label} =====\n')
        by=collections.defaultdict(list)
        for r in grp: by[r['category']].append(r['domain'])
        for c in sorted(by):
            o.write(f'# -- {c} ({len(by[c])}) --\n')
            for dd in sorted(by[c]): o.write(f'domain:{dd}\n')

# ---------- workbook ----------
wb=openpyxl.Workbook()
HEAD=Font(bold=True,color='FFFFFF'); HFILL=PatternFill('solid',fgColor='1F4E78')
THIN=Border(*[Side(style='thin',color='D9D9D9')]*4)
FILLS={'DISAVOW':PatternFill('solid',fgColor='F8CBAD'),'REVIEW':PatternFill('solid',fgColor='FFE699'),
       'KEEP':PatternFill('solid',fgColor='C6E0B4')}
CONF_FILL={'HIGH':PatternFill('solid',fgColor='F4B183'),'MEDIUM':PatternFill('solid',fgColor='FFD966'),
       'REVIEW':PatternFill('solid',fgColor='FFF2CC'),'KEEP':PatternFill('solid',fgColor='C6E0B4')}

def sheet(title,cols,data,widths=None,color_key=None):
    ws=wb.create_sheet(title)
    ws.append(cols)
    for i,c in enumerate(cols,1):
        cell=ws.cell(1,i); cell.font=HEAD; cell.fill=HFILL; cell.alignment=Alignment(vertical='center',wrap_text=True)
    for r in data:
        ws.append([r.get(k,'') for k in cols])
        if color_key:
            rc=ws.cell(ws.max_row, cols.index(color_key)+1)
            key=r.get(color_key)
            f=FILLS.get(key) or CONF_FILL.get(key)
            if f: rc.fill=f
    ws.freeze_panes='A2'
    ws.auto_filter.ref=f'A1:{get_column_letter(len(cols))}{ws.max_row}'
    widths=widths or {}
    for i,c in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(i)].width=widths.get(c,14)
    return ws

FULLCOLS=['domain','sources','decision','confidence','category','dr','ascore','traffic','positions',
          'is_spam','links','ip','country','first','last','topical','signals','evidence','qa_note']
W={'domain':30,'sources':13,'decision':11,'confidence':11,'category':26,'evidence':60,'qa_note':60,
   'ip':15,'category':26,'first':11,'last':11}

order={'DISAVOW':0,'REVIEW':1,'KEEP':2}
corder={'HIGH':0,'MEDIUM':1,'REVIEW':2,'KEEP':3}

# --- Summary ---
ws=wb.active; ws.title='Summary'
conf_c=collections.Counter(r['confidence'] for r in rows)
qa_c=sum(1 for r in rows if r['qa_note'] and 'CONFIRMED' not in r['qa_note'] and 'likely genuine' not in r['qa_note'])
summ=[
 ['Backlink Audit & Disavow — ngwindows.com',''],
 ['Generated',str(datetime.date.today())],
 ['Sources','Ahrefs (Site Explorer) + Semrush (Backlink Analytics)'],
 ['',''],
 ['Unique referring domains',len(all_domains)],
 ['  • Ahrefs-only',len(set(ah)-set(sm))],
 ['  • Semrush-only',len(set(sm)-set(ah))],
 ['  • In both (corroborated)',len(set(ah)&set(sm))],
 ['',''],
 ['DECISION BREAKDOWN',''],
 ['DISAVOW – HIGH confidence',conf_c['HIGH']],
 ['DISAVOW – MEDIUM confidence',conf_c['MEDIUM']],
 ['DISAVOW total (in disavow.txt)',conf_c['HIGH']+conf_c['MEDIUM']],
 ['REVIEW – manual (not auto-disavowed)',conf_c['REVIEW']],
 ['KEEP – protected',conf_c['KEEP']],
 ['',''],
 ['QA second-look items (see QA tab)',qa_c],
 ['',''],
 ['CONFIDENCE / DECISION LEGEND',''],
 ['HIGH','Self-evident spam footprint OR ≥2 independent toxic signals → disavow'],
 ['MEDIUM','Exactly one strong toxic signal → disavow after verify'],
 ['REVIEW','Low authority only, no hard footprint → manual call, NOT auto-disavowed'],
 ['KEEP','Recognized brand or authority ≥30 with no toxic footprint → do NOT disavow'],
 ['',''],
 ['TOXIC SIGNALS USED',''],
 ['1','Name footprint (backlink-seller, blogspot PBN, gambling, adult, shortener, stats/dir spam)'],
 ['2','Ahrefs is_spam=true'],
 ['3','Dedicated link-farm IP (≥4 of our refdomains share one non-CDN IP; CDN/cloud excluded)'],
 ['4','Known link-farm IP block / Moldova cluster'],
 ['5','Dead domain: DR/AS ≤3, 0 organic traffic, 0 ranking keywords'],
 ['',''],
 ['TABS','All Domains · Disavow-HIGH · Disavow-MEDIUM · Manual Review · Keep · QA Second-Look · Link Networks · Toxic Anchors · Ahrefs raw · Semrush raw'],
]
for r in summ: ws.append(r)
ws.column_dimensions['A'].width=38; ws.column_dimensions['B'].width=95
ws['A1'].font=Font(bold=True,size=14)
for rr in range(1,ws.max_row+1):
    lab=ws.cell(rr,1).value or ''
    if lab in ('DECISION BREAKDOWN','CONFIDENCE / DECISION LEGEND','TOXIC SIGNALS USED'):
        ws.cell(rr,1).font=Font(bold=True,color='1F4E78')

# --- All domains ---
alls=sorted(rows,key=lambda x:(order[x['decision']],corder[x['confidence']],x['category'],x['domain']))
sheet('All Domains',FULLCOLS,alls,W,color_key='decision')
# --- Disavow HIGH / MEDIUM ---
sheet('Disavow - HIGH',FULLCOLS,[r for r in alls if r['confidence']=='HIGH'],W,color_key='confidence')
sheet('Disavow - MEDIUM',FULLCOLS,[r for r in alls if r['confidence']=='MEDIUM'],W,color_key='confidence')
# --- Manual Review ---
sheet('Manual Review',FULLCOLS,[r for r in alls if r['confidence']=='REVIEW'],W,color_key='confidence')
# --- Keep ---
sheet('Keep (Protected)',FULLCOLS,[r for r in alls if r['confidence']=='KEEP'],W,color_key='confidence')
# --- QA second look ---
qarows=[r for r in alls if r['qa_note']]
sheet('QA Second-Look',['domain','decision','confidence','category','dr','ascore','traffic','positions','is_spam','ip','qa_note'],
      qarows,{'domain':30,'category':26,'qa_note':85,'ip':15},color_key='decision')
# --- Link networks ---
netcols=['ip','domains_on_ip','type','member_domains']
netrows=[]
for ip,doms in sorted(network_ips.items(),key=lambda x:-len(x[1])):
    typ='Known link-farm' if ip in FARM_IPS else ('Moldova cluster' if ip==MOLDOVA_IP else 'Shared non-CDN host')
    netrows.append(dict(ip=ip,domains_on_ip=len(doms),type=typ,member_domains=', '.join(sorted(doms))))
sheet('Link Networks',netcols,netrows,{'ip':16,'domains_on_ip':14,'type':20,'member_domains':120})
# --- Toxic anchors ---
anchors=[
 ('...SEOExpress.org and their backlink building service truly worked wonders! ...traffic increased by over 400%',222,'TRUE'),
 ('Complete SEO for ngwindows.com: premium guest posts, contextual backlinks, on-page and local SEO... DR/DA/TF gains',44,'TRUE'),
 ('High Quality Dofollow Backlinks DA 50 PA 40 Premium PBN Network Service ngwindows.com Rank First Page Google Fast SEO Link Building Buy Backlinks Online Cheap',21,'TRUE'),
 ('JOIN OUR TELEGRAM https://t.me/s/darksidelinks',10,'TRUE'),
 ('OUR TELEGRAM CHANEL https://t.me/s/quarterlinks25',9,'FALSE'),
 ('Complete SEO for ngwindows.com: guest posts and backlinks, on-page SEO, local SEO, web development...',5,'TRUE'),
 ('Get free high quality HD wallpapers printable psychrometric chart (scraper/wallpaper spam)',15,'FALSE'),
 ('tLvtiPx5V2OVDzj (random gibberish anchor)',11,'TRUE'),
]
sheet('Toxic Anchors',['anchor_text','referring_domains','ahrefs_is_spam'],
      [dict(anchor_text=a,referring_domains=n,ahrefs_is_spam=sp) for a,n,sp in anchors],
      {'anchor_text':110,'referring_domains':18,'ahrefs_is_spam':16})
# --- raw ---
sheet('Ahrefs raw',['domain','dr','traffic','positions','is_spam','links','ip','first','last'],
      [dict(domain=d,dr=round(v['dr'],1),traffic=v['traffic'],positions=v['positions'],
            is_spam='TRUE' if v['is_spam'] else '',links=v['links'],ip=v['ip'],first=v['first'],last=v['last'])
       for d,v in sorted(ah.items())],{'domain':32})
sheet('Semrush raw',['domain','ascore','trust','links','ip','country','first','last'],
      [dict(domain=d,ascore=v['ascore'],trust=v['trust'],links=v['links'],ip=v['ip'],country=v['country'],
            first=v['first'],last=v['last']) for d,v in sorted(sm.items())],{'domain':32})

wb.save('ngwindows_backlink_audit.xlsx')
print('workbook saved: ngwindows_backlink_audit.xlsx')
print('sheets:',wb.sheetnames)
print('decisions:',dict(collections.Counter(r['decision'] for r in rows)))
print('confidence:',dict(conf_c))
print('QA second-look rows:',len(qarows))
# also refresh flagged CSV
FCOLS=FULLCOLS
with open('disavow_flagged_for_review.csv','w',newline='') as o:
    w=csv.DictWriter(o,fieldnames=FCOLS); w.writeheader()
    for r in alls:
        if r['confidence']!='KEEP': w.writerow(r)
with open('merged_all_domains.csv','w',newline='') as o:
    w=csv.DictWriter(o,fieldnames=FCOLS); w.writeheader()
    for r in alls: w.writerow(r)
print('refreshed CSVs')
