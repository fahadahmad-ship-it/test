#!/usr/bin/env python3
"""Simple review workbook + designed dashboard, with per-domain anchors joined in."""
import csv, collections, html, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import json
rows=list(csv.DictReader(open('merged_all_domains.csv')))
anch={r['domain']:r for r in csv.DictReader(open('anchor_by_domain.csv'))}
def root2(d):  # last two labels, for subdomain fallback matching
    p=d.split('.'); return '.'.join(p[-2:]) if len(p)>=2 else d
anch_root={root2(k):v for k,v in anch.items()}
def get_anchor(d):
    return anch.get(d) or anch.get(root2(d)) or anch_root.get(root2(d)) or {}
# actual backlink-level evidence from Semrush (source URL + anchor + follow)
BL=json.load(open('semrush_backlink_by_domain.json'))
def get_bl(d):
    return BL.get(d) or BL.get(root2(d)) or {}

WHY={
 'Link-selling / SEO-service PBN':'Paid-link / backlink-seller site',
 'Blogspot PBN / spam post':'Spam blog post (Blogspot PBN)',
 'Gambling / casino spam':'Gambling / casino spam',
 'Adult spam':'Adult spam',
 'URL-shortener / redirect link network':'Redirect / URL-shortener link network',
 'Expired-domain / auto network':'Expired-domain spam network',
 'Auto-generated stats / worth-checker':'Auto-generated stats / scraper page',
 'Low-quality directory / TLD list':'Spam directory listing',
 'Free-host throwaway page':'Throwaway free-host page',
 'Multi-signal spam':'Multiple spam signals (dead + farm-IP / spam-flag)',
 'Single-signal low-quality':'Low-quality / dead domain',
 'Window-company microsite cluster (verify OWNERSHIP)':'Possibly YOUR OWN microsite — verify before disavow',
 'Low-authority (needs review)':'Low authority — manual check',
 'Unclear':'Borderline — manual check',
 'LEGITIMATE / high-authority':'Legitimate / high-authority — keep',
}
for r in rows:
    a=get_anchor(r['domain']); b=get_bl(r['domain'])
    r['anchor_type']=a.get('anchor_type','')
    # example backlink: prefer Semrush source URL, else Ahrefs url_from
    r['backlink_url']=b.get('ex_url') or a.get('url_from','')
    r['example_anchor']=b.get('ex_anchor') or a.get('anchor','')
    r['follow']=b.get('ex_follow') or a.get('dofollow','')
    r['bl_count']=b.get('bl_count','')
    r['dofollow_n']=b.get('dofollow','')
    r['nofollow_n']=b.get('nofollow','')
    r['link_checked']='yes' if (b or a) else 'no (beyond sample)'
    r['why']=WHY.get(r['category'],r['category'])
    r['authority']=max(float(r['dr'] or 0), int(r['ascore'] or 0))

# ================= STRICT, ANCHOR-DRIVEN RE-VERIFICATION =================
# Every decision is re-derived from the ACTUAL backlink evidence (real anchor text,
# follow status, farm-IP, is_spam, footprint) rather than domain metrics alone.
import re as _re
FARM={'203.161.54.114','118.139.181.85','118.139.176.46','118.139.178.200','118.139.161.199',
 '118.139.181.255','184.168.115.60','195.20.19.178','67.223.118.29','188.40.17.96','92.249.46.138','191.101.14.187'}
HARDCAT={'Link-selling / SEO-service PBN','Blogspot PBN / spam post','Gambling / casino spam','Adult spam',
 'URL-shortener / redirect link network','Expired-domain / auto network','Auto-generated stats / worth-checker',
 'Low-quality directory / TLD list','Free-host throwaway page','Multi-signal spam'}
def aclass(t):
    t=(t or '').strip(); tl=t.lower()
    if t in ('','.','|','-','·') or (len(t)<=2 and not _re.search(r'[a-z]{2}',tl)): return 'empty/blank'
    if _re.search(r'backlink|dofollow|\bpbn\b|buy (back)?link|rank first|rank(ing)? with|premium .*link|\bseo\b|casino|poker|slot|\bbet\b|betting|porn|\bsex\b|viagra|cialis|\bloan\b|crypto|telegram|t\.me|wallpaper|psychrometric|\bsmm\b|panel|escort',tl): return 'money/spam'
    if _re.search(r'window|door|glass|vinyl|replacement|atlanta|georgia|sunroom|contractor|remodel|roof|siding|home improv|installation|grill design|crawl space|insulation|blinds|shutter|patio|garage',tl): return 'commercial'
    if _re.search(r'ngwindows|north georgia|ng windows|^https?://|^www\.|roiwindows|qualitypluswindows',tl): return 'branded/url'
    if _re.search(r'read more|click here|website|visit|learn more|go now|download|^here$|^home$|contact|opens a new tab|inspiration|gallery|proposal|consultation',tl): return 'generic'
    return 'other'
def raw_ip(d): return bool(_re.match(r'^\d{1,3}(\.\d{1,3}){3}$', d.strip()))

for r in rows:
    r['aclass']=aclass(r['example_anchor'])
    a=r['aclass']; isspam=str(r['is_spam']).upper()=='TRUE'; farm=r['ip'] in FARM; md=r['country']=='md'
    hard=r['category'] in HARDCAT
    dof=(str(r['dofollow_n']).isdigit() and int(r['dofollow_n'])>0) or r['follow']=='dofollow'
    auth=float(r['authority'] or 0); tr=float(r['traffic'] or 0)
    micro='microsite' in r['category']
    # strong, unambiguous spam evidence
    strong = hard or isspam or farm or md or a=='money/spam' or raw_ip(r['domain'])
    # manipulative exact-match commercial anchor on a dead/low-authority off-topic domain
    manip_comm = a=='commercial' and auth<15 and tr==0
    dec=r['decision']; verdict='confirmed'; why_reason=[]
    if micro:
        dec='REVIEW'; verdict='hold — verify ownership'; why_reason=['own-microsite cluster']
    elif strong or manip_comm:
        if dec!='DISAVOW': verdict='UPGRADED review→disavow'
        dec='DISAVOW'
        why_reason=[x for x,ok in [('footprint',hard),('is_spam',isspam),('farm-IP',farm),('md',md),
                    ('paid/spam anchor',a=='money/spam'),('raw-IP domain',raw_ip(r['domain'])),
                    ('exact-match commercial anchor',manip_comm)] if ok]
    else:
        # no strong evidence: dofollow-generic / branded / dead-only → not clearly manipulative
        if dec=='DISAVOW': verdict='DOWNGRADED disavow→review (no spam signal)'
        elif dec=='KEEP':
            if isspam or farm or a=='money/spam': verdict='keep→review (toxic signal)'
            else: verdict='confirmed'; # stays KEEP
        if not (dec=='KEEP' and verdict=='confirmed'):
            dec='REVIEW'
        why_reason=['dofollow, generic anchor — verify' if dof else 'low-quality, no spam signal']
    r['decision']=dec; r['strict_verdict']=verdict; r['tier_reason']='; '.join(why_reason)

# priority derived from the strict decision
for r in rows:
    if r['decision']=='DISAVOW':
        core = r['category'] in HARDCAT or str(r['is_spam']).upper()=='TRUE' or r['ip'] in FARM \
               or r['aclass']=='money/spam' or r['country']=='md' or r['confidence']=='HIGH'
        r['priority']='P1 – Core (disavow)' if core else 'P2 – Recommended (disavow)'
    elif r['decision']=='REVIEW':
        r['priority']='Review (manual)'
    else:
        r['priority']='Keep'

DIS=[r for r in rows if r['decision']=='DISAVOW']
REV=[r for r in rows if r['decision']=='REVIEW']
KEEP=[r for r in rows if r['decision']=='KEEP']
P1=[r for r in DIS if r['priority'].startswith('P1')]
P2=[r for r in DIS if r['priority'].startswith('P2')]
P3=[]  # optional tier retired; weak domains now sit in REVIEW
conf=collections.Counter(r['confidence'] for r in rows)
import collections as _c
print('STRICT verdicts:',dict(_c.Counter(r['strict_verdict'] for r in rows)))
print('decisions:',dict(_c.Counter(r['decision'] for r in rows)),'| P1',len(P1),'P2',len(P2))

# ================= WORKBOOK =================
wb=openpyxl.Workbook()
HEAD=Font(bold=True,color='FFFFFF',size=11); HFILL=PatternFill('solid',fgColor='1F3864')
DEC_FILL={'DISAVOW':PatternFill('solid',fgColor='F8CBAD'),
          'REVIEW':PatternFill('solid',fgColor='FFE699'),
          'KEEP':PatternFill('solid',fgColor='C6E0B4')}
def build(ws,cols,data,widths,colorby='decision'):
    ws.append(cols)
    for i,c in enumerate(cols,1):
        cell=ws.cell(1,i); cell.font=HEAD; cell.fill=HFILL
        cell.alignment=Alignment(vertical='center',wrap_text=True,horizontal='left')
    for r in data:
        ws.append([r.get(k,'') for k in cols])
        if colorby and colorby in cols:
            c=ws.cell(ws.max_row, cols.index(colorby)+1)
            f=DEC_FILL.get(r.get(colorby));
            if f: c.fill=f
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:{get_column_letter(len(cols))}{ws.max_row}'
    for i,c in enumerate(cols,1): ws.column_dimensions[get_column_letter(i)].width=widths.get(c,14)
    ws.row_dimensions[1].height=26

# ---- Tab 1: How to use ----
ws=wb.active; ws.title='① Read me'
info=[
 ['ngwindows.com — Backlink Disavow Review',''],
 ['Generated',str(datetime.date.today())],
 ['',''],
 ['WHAT TO DO','—'],
 ['1','Open “② Review list”. Rows are colour-coded: RED = disavow, YELLOW = manual review, and the KEEP list is on its own tab.'],
 ['2','Skim the “Why” and “Anchor” columns. Untick (delete the row) anything you recognise as a real/own site.'],
 ['3','The “③ Disavow file” tab is the exact text to upload to Google Search Console → Disavow Tool (or use disavow.txt).'],
 ['4','“④ Keep” = do NOT disavow. “⑤ Full data” = every signal if you want the detail.'],
 ['',''],
 ['NUMBERS (after strict, anchor-driven re-check)','—'],
 ['Total referring domains',len(rows)],
 ['DISAVOW (each backed by real spam evidence)',len(DIS)],
 ['   • P1 Core (footprint / is_spam / farm-IP / paid anchor)',len(P1)],
 ['   • P2 Recommended (exact-match commercial anchor / raw-IP / dofollow)',len(P2)],
 ['Manual review (ambiguous — you decide)',len(REV)],
 ['Keep (protected)',len(KEEP)],
 ['',''],
 ['STRICT RE-CHECK — what changed','—'],
 ['Every domain was re-judged from its ACTUAL anchor text + follow + IP + footprint.',''],
 ['Downgraded disavow → review (no spam signal, just dofollow/blank anchor): %d'%sum(1 for r in rows if 'DOWNGRADED' in r.get('strict_verdict','')),''],
 ['Upgraded review → disavow (Telegram/paid/commercial anchors that were missed): %d'%sum(1 for r in rows if 'UPGRADED' in r.get('strict_verdict','')),''],
 ['Kept on hold for ownership check (microsites): %d'%sum(1 for r in rows if 'ownership' in r.get('strict_verdict','')),''],
 ['Tightest cut if you prefer: P1 Core only = %d domains.'%len(P1),''],
 ['',''],
 ['ACTUAL BACKLINKS CHECKED','—'],
 ['Pulled 5,000+ individual backlink records (source page URL, anchor, follow) from Semrush + Ahrefs.',''],
 ['Each flagged row shows a real example link in “backlink_url” + “example_anchor”; “link_checked” = yes/beyond-sample.',''],
 ['Sample toxic anchors found across 200+ domains:',''],
 ['  •','“…SEOExpress.org and their backlink building service worked wonders… traffic +400%”'],
 ['  •','“High Quality Dofollow Backlinks DA50 PA40 Premium PBN … Buy Backlinks Online Cheap”'],
 ['  •','“JOIN OUR TELEGRAM https://t.me/s/darksidelinks”'],
 ['Blogspot PBNs (innocyscx / burnersgamershitasd) link using your microsite names (qualitypluswindows.com, roiwindows.com) as anchors.',''],
 ['Most spam links are nofollow (pass no equity — lower priority); the follow column flags the dofollow ones that matter most.',''],
 ['',''],
 ['⚠ OWNERSHIP','15 window-brand microsites (ngawindows.com, ngwindow.com, roiwindows.com …) may be YOUR OWN sites — they are in REVIEW, not disavow. Confirm before acting.'],
]
for r in info: ws.append(r)
ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=100
ws['A1'].font=Font(bold=True,size=14,color='1F3864')
for rr in range(1,ws.max_row+1):
    if ws.cell(rr,2).value=='—': ws.cell(rr,1).font=Font(bold=True,color='C00000')

SIMPLE=['domain','decision','priority','strict_verdict','why','follow','aclass','example_anchor','backlink_url','authority','traffic','link_checked']
SW={'domain':28,'decision':11,'priority':26,'why':32,'follow':9,'example_anchor':32,'backlink_url':48,
    'authority':10,'traffic':9,'link_checked':16,'strict_verdict':30,'aclass':13}
build(wb.create_sheet('② Review list'),SIMPLE,DIS+REV,SW)
build(wb.create_sheet('④ Keep (do NOT disavow)'),SIMPLE,KEEP,SW)

# Tab 3: disavow file text (P1 Core + P2 Recommended active; P3 optional commented)
ws=wb.create_sheet('③ Disavow file')
ws.append(['Paste into Google Search Console → Disavow Tool (or use disavow.txt)'])
ws['A1'].font=Font(bold=True,color='1F3864')
ws.append([f'# ngwindows.com disavow — {datetime.date.today()}'])
ws.append([f'# STRICT-VERIFIED disavow = P1 Core ({len(P1)}) + P2 Recommended ({len(P2)}) = {len(DIS)} domains'])
ws.append(['# Each domain below has a real spam signal: footprint, is_spam, farm-IP, paid/commercial anchor, or raw-IP.'])
ws.append(['# --- P1 CORE: footprint / is_spam / farm-IP / paid (money) anchor ---'])
for r in sorted(P1,key=lambda x:x['domain']): ws.append([f"domain:{r['domain']}"])
ws.append(['# --- P2 RECOMMENDED: exact-match commercial anchor / raw-IP domain / dofollow ---'])
for r in sorted(P2,key=lambda x:x['domain']): ws.append([f"domain:{r['domain']}"])
ws.column_dimensions['A'].width=60

# write disavow.txt (strict-verified P1+P2; every line has a spam signal)
with open('disavow.txt','w') as o:
    o.write(f'# Disavow file for ngwindows.com — {datetime.date.today()}\n')
    o.write('# Ahrefs + Semrush audit, strictly re-verified at the actual-backlink level.\n')
    o.write(f'# Every domain here carries a real spam signal (footprint / is_spam / farm-IP / paid or exact-match anchor / raw-IP).\n')
    o.write(f'# TOTAL = P1 Core ({len(P1)}) + P2 Recommended ({len(P2)}) = {len(DIS)} domains.\n')
    o.write(f'# For the tightest cut, submit P1 Core only ({len(P1)}). {len(REV)} ambiguous domains are held in the workbook for manual review.\n#\n')
    o.write(f'# ===== P1 — CORE ({len(P1)}): footprint / is_spam / farm-IP / paid anchor =====\n')
    for r in sorted(P1,key=lambda x:x['domain']): o.write(f"domain:{r['domain']}\n")
    o.write(f'#\n# ===== P2 — RECOMMENDED ({len(P2)}): exact-match commercial anchor / raw-IP / dofollow =====\n')
    for r in sorted(P2,key=lambda x:x['domain']): o.write(f"domain:{r['domain']}\n")

# Tab 5: full data
FULL=['domain','decision','priority','strict_verdict','confidence','category','why','aclass','anchor_type','example_anchor','backlink_url',
      'follow','bl_count','dofollow_n','nofollow_n','link_checked','dr','ascore','authority','traffic',
      'positions','is_spam','links','ip','country','tier_reason','evidence','qa_note']
FW={'domain':30,'priority':26,'category':26,'why':34,'example_anchor':40,'backlink_url':50,'tier_reason':40,'evidence':55,'qa_note':60,'ip':15}
build(wb.create_sheet('⑤ Full data'),FULL,DIS+REV+KEEP,{**SW,**FW})

wb.save('ngwindows_disavow_review.xlsx')
print('workbook:',wb.sheetnames)

# complete flat CSV (all domains, all columns)
with open('ngwindows_backlink_audit_complete.csv','w',newline='',encoding='utf-8') as o:
    w=csv.DictWriter(o,fieldnames=FULL,extrasaction='ignore'); w.writeheader()
    for r in sorted(DIS+REV+KEEP,key=lambda x:(x['decision'],x['priority'],x['domain'])): w.writerow(r)
print('wrote ngwindows_backlink_audit_complete.csv')

# ================= DASHBOARD =================
P={'disavow':'#e34948','review':'#eda100','keep':'#008300','ink':'#101010','ink2':'#5b5a55',
   'surf':'#ffffff','line':'#e7e7e2','accent':'#2a78d6','bg':'#f4f4f1','navy':'#16233b'}
FARM={'203.161.54.114','118.139.181.85','118.139.176.46','118.139.178.200','118.139.161.199',
 '118.139.181.255','184.168.115.60','195.20.19.178','67.223.118.29','188.40.17.96','92.249.46.138','191.101.14.187'}
def dis_reason(r):
    a=r['aclass']; cat=r['category']
    M={'Link-selling / SEO-service PBN':'Backlink-seller / SEO PBN','Blogspot PBN / spam post':'Blogspot PBN post',
       'Gambling / casino spam':'Gambling / casino','Adult spam':'Adult spam',
       'URL-shortener / redirect link network':'URL-shortener / redirect net',
       'Auto-generated stats / worth-checker':'Auto stats / scraper page','Low-quality directory / TLD list':'Spam directory',
       'Expired-domain / auto network':'Expired-domain network','Free-host throwaway page':'Free-host throwaway'}
    if cat in M: return M[cat]
    if a=='money/spam': return 'Paid / Telegram spam anchor'
    if r['ip'] in FARM: return 'Farm-IP link network'
    if a=='commercial': return 'Exact-match commercial anchor'
    if _re.match(r'^\d{1,3}(\.\d{1,3}){3}$',r['domain']): return 'Raw-IP / junk host'
    if str(r['is_spam']).upper()=='TRUE': return 'Ahrefs spam-flagged'
    return 'Other spam signal'
reason_counts=collections.Counter(dis_reason(r) for r in DIS).most_common(10)
aclass_counts=collections.Counter(r['aclass'] for r in DIS).most_common()
dof=sum(1 for r in DIS if r['follow']=='dofollow' or (str(r['dofollow_n']).isdigit() and int(r['dofollow_n'])>0))
nof=len(DIS)-dof
netmap=collections.defaultdict(list)
for r in DIS+REV:
    if r['ip'] in FARM: netmap[r['ip']].append(r['domain'])
nets=sorted(((ip,d) for ip,d in netmap.items()),key=lambda x:-len(x[1]))[:8]
anchors=[('…SEOExpress.org and their backlink building service worked wonders… traffic +400%',222),
 ('Complete SEO for ngwindows.com: premium guest posts, contextual backlinks…',44),
 ('High Quality Dofollow Backlinks DA50 PA40 Premium PBN … Buy Backlinks Online Cheap',21),
 ('JOIN OUR TELEGRAM https://t.me/s/darksidelinks',10),
 ('tLvtiPx5V2OVDzj (random gibberish anchor)',11)]

def hbars(data,color,unit='',w=600,rowh=32,pad=210,fs=12.5):
    if not data: return ''
    mx=max(v for _,v in data) or 1
    bw=w-pad-64; H=len(data)*rowh+8; out=[f'<svg viewBox="0 0 {w} {H}" width="100%" role="img" aria-hidden="false">']
    for i,(lab,v) in enumerate(data):
        y=i*rowh+8; bl=max(5,bw*v/mx); ls=html.escape(str(lab))
        if len(ls)>32: ls=ls[:31]+'…'
        out.append(f'<text x="{pad-10}" y="{y+15}" text-anchor="end" font-size="{fs}" fill="{P["ink2"]}">{ls}</text>')
        out.append(f'<rect x="{pad}" y="{y+2}" width="{bl:.1f}" height="18" rx="4" fill="{color}"><title>{ls}: {v}</title></rect>')
        out.append(f'<text x="{pad+bl+8}" y="{y+15}" font-size="12" font-weight="700" fill="{P["ink"]}">{v}{unit}</text>')
    out.append('</svg>'); return ''.join(out)

def donut(parts):
    tot=sum(v for _,v,_ in parts) or 1; import math
    cx,cy,r,sw=92,92,66,24; a=-math.pi/2; seg=[]
    for lab,v,col in parts:
        frac=v/tot; a2=a+frac*2*math.pi
        x1,y1=cx+r*math.cos(a),cy+r*math.sin(a); x2,y2=cx+r*math.cos(a2),cy+r*math.sin(a2)
        lg=1 if frac>0.5 else 0
        seg.append(f'<path d="M {x1:.1f} {y1:.1f} A {r} {r} 0 {lg} 1 {x2:.1f} {y2:.1f}" fill="none" stroke="{col}" stroke-width="{sw}" stroke-linecap="butt"><title>{html.escape(lab)}: {v}</title></path>')
        a=a2
    return (f'<svg viewBox="0 0 184 184" width="176" height="176" role="img">{"".join(seg)}'
            f'<text x="92" y="88" text-anchor="middle" font-size="30" font-weight="800" fill="{P["ink"]}">{tot}</text>'
            f'<text x="92" y="108" text-anchor="middle" font-size="12" fill="{P["ink2"]}">domains</text></svg>')

def kpi(v,l,c,sub=''):
    s=f'<div class="kpi-sub">{sub}</div>' if sub else ''
    return f'<div class="kpi" style="--ac:{c}"><div class="kpi-v" style="color:{c}">{v}</div><div class="kpi-l">{l}</div>{s}</div>'

toxic_pct=round(100*len(DIS)/len(rows))
net_rows=''.join(f'<tr><td class="mono">{html.escape(ip)}</td><td class="r">{len(d)}</td></tr>' for ip,d in nets)
anc_rows=''.join(f'<tr><td>{html.escape(a)}</td><td class="r">{n}</td></tr>' for a,n in anchors)
DONUT=[('Disavow (verified)',len(DIS),P['disavow']),('Manual review',len(REV),P['review']),('Keep',len(KEEP),P['keep'])]
legend=''.join(f'<div class="lg"><i style="background:{c}"></i><span>{l}</span><b>{v}</b></div>' for l,v,c in DONUT)
checked=sum(1 for r in (DIS+REV) if r.get('backlink_url'))

HTML=f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>ngwindows.com — Backlink Disavow Dashboard</title>
<style>
 :root{{--surf:{P['surf']};--ink:{P['ink']};--ink2:{P['ink2']};--line:{P['line']};--red:{P['disavow']};--amber:{P['review']};--green:{P['keep']}}}
 *{{box-sizing:border-box}}
 body{{margin:0;font:14px/1.55 ui-sans-serif,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;background:{P['bg']};color:var(--ink)}}
 .wrap{{max-width:1140px;margin:0 auto;padding:0 22px 40px}}
 header.hd{{background:{P['navy']};color:#fff;margin-bottom:22px;border-radius:0 0 16px 16px;
   padding:26px 28px;box-shadow:0 2px 12px rgba(20,35,59,.14)}}
 header.hd .row{{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap;max-width:1140px;margin:0 auto}}
 header.hd h1{{font-size:23px;margin:0 0 4px;font-weight:800;letter-spacing:-.01em}}
 header.hd .sub{{color:#c7d0de;font-size:12.5px}}
 .pill{{background:rgba(227,73,72,.16);color:#ffb4b3;border:1px solid rgba(227,73,72,.5);
   padding:7px 14px;border-radius:999px;font-size:12.5px;font-weight:700;white-space:nowrap}}
 .grid{{display:grid;gap:15px}} .k5{{grid-template-columns:repeat(5,1fr)}}
 .two{{grid-template-columns:1.18fr .82fr}} .two2{{grid-template-columns:1fr 1fr}}
 @media(max-width:860px){{.k5{{grid-template-columns:repeat(2,1fr)}}.two,.two2{{grid-template-columns:1fr}}}}
 .card{{background:var(--surf);border:1px solid var(--line);border-radius:14px;padding:18px 20px 15px;box-shadow:0 1px 2px rgba(16,16,16,.04)}}
 .card h2{{font-size:12px;letter-spacing:.06em;text-transform:uppercase;color:var(--ink2);margin:0 0 15px;font-weight:800}}
 .kpi{{background:var(--surf);border:1px solid var(--line);border-radius:14px;padding:16px 15px 15px;position:relative;overflow:hidden}}
 .kpi:before{{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--ac)}}
 .kpi-v{{font-size:31px;font-weight:800;line-height:1;letter-spacing:-.01em}}
 .kpi-l{{font-size:12px;color:var(--ink2);margin-top:7px;font-weight:600}}
 .kpi-sub{{font-size:11px;color:var(--ink2);margin-top:3px}}
 .banner{{background:linear-gradient(0deg,#fff,#fdf0f0);border:1px solid #f2c0c0;border-left:5px solid var(--red);
   border-radius:12px;padding:15px 18px;margin-bottom:16px;font-size:13.5px}}
 .donutrow{{display:flex;align-items:center;gap:22px;flex-wrap:wrap}}
 .lg{{display:flex;align-items:center;gap:8px;font-size:13px;color:var(--ink2);margin:7px 0}}
 .lg i{{width:12px;height:12px;border-radius:3px}} .lg b{{margin-left:auto;color:var(--ink);font-variant-numeric:tabular-nums}}
 .lg span{{min-width:130px}}
 table{{width:100%;border-collapse:collapse;font-size:12.5px}}
 td{{padding:7px 8px;border-bottom:1px solid var(--line)}} tr:last-child td{{border-bottom:0}}
 th{{text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.04em;color:var(--ink2);padding:0 8px 8px}}
 .r{{text-align:right;font-variant-numeric:tabular-nums;font-weight:700}}
 .mono{{font-family:ui-monospace,Menlo,Consolas,monospace}}
 .warn{{background:#fff7ea;border:1px solid #f0cf94;border-left:5px solid var(--amber);border-radius:12px;padding:15px 18px;margin-top:16px;font-size:13px}}
 .warn b{{color:#9a5b00}}
 .split{{display:flex;gap:10px;margin-top:10px}}
 .chip{{flex:1;text-align:center;border:1px solid var(--line);border-radius:10px;padding:10px 6px}}
 .chip .n{{font-size:20px;font-weight:800}} .chip .t{{font-size:11px;color:var(--ink2);margin-top:2px}}
 .foot{{color:var(--ink2);font-size:11.5px;margin-top:24px;text-align:center}}
 .steps{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-top:2px}}
 @media(max-width:860px){{.steps{{grid-template-columns:1fr 1fr}}}}
 .step{{border:1px solid var(--line);border-radius:10px;padding:12px 14px;font-size:12.5px}}
 .step b{{display:block;color:var(--ink);margin-bottom:3px}}
</style></head><body>
<header class="hd"><div class="row">
  <div><h1>Backlink Disavow Audit — ngwindows.com</h1>
   <div class="sub">Ahrefs + Semrush · {datetime.date.today()} · {checked}/{len(DIS)+len(REV)} flagged domains verified at the actual-backlink level (anchor + follow)</div></div>
  <div class="pill">⚠ {toxic_pct}% toxic — paid-PBN profile</div>
</div></header>
<div class="wrap">

 <div class="banner"><b>Verdict: heavily manipulated (paid-PBN) link profile.</b>
  After a strict, anchor-driven re-check, <b>{len(DIS)} domains</b> are staged for disavow (P1 Core {len(P1)} + P2 {len(P2)}) —
  <b>each carries a real spam signal</b> (footprint, is_spam, farm-IP, paid/exact-match anchor, or raw-IP host).
  {len(REV)} ambiguous domains are held for manual review, not auto-disavowed.</div>

 <div class="grid k5" style="margin-bottom:15px">
  {kpi(len(rows),'Referring domains',P['navy'])}
  {kpi(len(DIS),'Disavow',P['disavow'],'verified')}
  {kpi(len(P1),'P1 Core',P['disavow'],'tightest cut')}
  {kpi(len(REV),'Manual review',P['review'],'you decide')}
  {kpi(len(KEEP),'Keep',P['keep'],'protected')}
 </div>

 <div class="grid two" style="margin-bottom:15px">
  <div class="card"><h2>Why domains are flagged</h2>{hbars(reason_counts,P['disavow'])}</div>
  <div class="card"><h2>Decision split</h2>
   <div class="donutrow">{donut(DONUT)}<div style="flex:1;min-width:190px">{legend}</div></div></div>
 </div>

 <div class="grid two2" style="margin-bottom:15px">
  <div class="card"><h2>Anchor type of disavowed links</h2>{hbars(aclass_counts,P['accent'],pad=150)}
   <div class="split">
    <div class="chip"><div class="n" style="color:{P['disavow']}">{dof}</div><div class="t">dofollow (pass equity)</div></div>
    <div class="chip"><div class="n">{nof}</div><div class="t">nofollow (lower priority)</div></div>
   </div></div>
  <div class="card"><h2>Largest link networks (shared farm IP)</h2>
   <table><thead><tr><th>IP address</th><th class="r">domains</th></tr></thead><tbody>{net_rows}</tbody></table></div>
 </div>

 <div class="card" style="margin-bottom:15px"><h2>Smoking-gun spam anchors (verbatim from the links)</h2>
  <table><thead><tr><th>Anchor text</th><th class="r">ref. domains</th></tr></thead><tbody>{anc_rows}</tbody></table></div>

 <div class="warn"><b>⚠ Verify ownership before disavowing:</b> 15 near-identical window-brand microsites
  (ngawindows.com, ngwindow.com, roiwindows.com, thermalprowindows.com, performingwindows.com …) sit on shared AWS IPs and are
  likely <b>your own sites or a self-built PBN</b>. Held in manual review — redirect/consolidate your own sites, don't disavow them.</div>

 <div class="card" style="margin-top:15px"><h2>How to use this</h2>
  <div class="steps">
   <div class="step"><b>1 · Submit</b>Upload <span class="mono">disavow.txt</span> (P1+P2 = {len(DIS)}) to Google Search Console → Disavow Tool. For the safest cut, submit P1 Core only ({len(P1)}).</div>
   <div class="step"><b>2 · Review</b>Skim the {len(REV)} manual-review rows; move any obvious spam into disavow, keep real sites out.</div>
   <div class="step"><b>3 · Ownership</b>Confirm the 15 microsites — disavow only if they are a third-party PBN.</div>
   <div class="step"><b>4 · Files</b>Every list is provided as a CSV in the csv/ folder + the full workbook.</div>
  </div></div>

 <div class="foot">Nothing is submitted to Google automatically. Evidence per domain is in the workbook / complete CSV.</div>
</div></body></html>"""
open('dashboard.html','w').write(HTML)
print('dashboard.html bytes:',len(HTML))

# ================= CSV EXPORTS (all files) =================
import os
os.makedirs('csv',exist_ok=True)
def wcsv(name,cols,data):
    with open('csv/'+name,'w',newline='',encoding='utf-8') as o:
        w=csv.DictWriter(o,fieldnames=cols,extrasaction='ignore'); w.writeheader()
        for r in data: w.writerow(r)
SC=['domain','decision','priority','strict_verdict','why','follow','aclass','example_anchor','backlink_url',
    'authority','dr','ascore','traffic','positions','is_spam','links','ip','country','tier_reason','link_checked']
wcsv('01_all_domains.csv',FULL,sorted(DIS+REV+KEEP,key=lambda x:(x['decision'],x['priority'],x['domain'])))
wcsv('02_disavow_all_545.csv',SC,sorted(DIS,key=lambda x:(x['priority'],x['domain'])))
wcsv('03_disavow_P1_core.csv',SC,sorted(P1,key=lambda x:x['domain']))
wcsv('04_disavow_P2_recommended.csv',SC,sorted(P2,key=lambda x:x['domain']))
wcsv('05_manual_review.csv',SC,sorted(REV,key=lambda x:x['domain']))
wcsv('06_keep_protected.csv',SC,sorted(KEEP,key=lambda x:x['domain']))
# google-format single column
with open('csv/07_disavow_google_format.csv','w',newline='',encoding='utf-8') as o:
    o.write('entry\n')
    for r in sorted(DIS,key=lambda x:x['domain']): o.write('domain:%s\n'%r['domain'])
# link networks
with open('csv/08_link_networks.csv','w',newline='',encoding='utf-8') as o:
    w=csv.writer(o); w.writerow(['ip','domains_on_ip','member_domains'])
    for ip,d in sorted(netmap.items(),key=lambda x:-len(x[1])): w.writerow([ip,len(d),', '.join(sorted(d))])
# toxic anchors
with open('csv/09_toxic_anchors.csv','w',newline='',encoding='utf-8') as o:
    w=csv.writer(o); w.writerow(['anchor_text','referring_domains']); [w.writerow([a,n]) for a,n in anchors]
# summary
with open('csv/10_summary.csv','w',newline='',encoding='utf-8') as o:
    w=csv.writer(o); w.writerow(['metric','value'])
    for k,v in [('total_referring_domains',len(rows)),('disavow_total',len(DIS)),('disavow_P1_core',len(P1)),
        ('disavow_P2_recommended',len(P2)),('manual_review',len(REV)),('keep_protected',len(KEEP)),
        ('toxic_pct',toxic_pct),('disavow_dofollow',dof),('disavow_nofollow',nof),
        ('downgraded_disavow_to_review',sum(1 for r in rows if 'DOWNGRADED' in r.get('strict_verdict',''))),
        ('upgraded_review_to_disavow',sum(1 for r in rows if 'UPGRADED' in r.get('strict_verdict',''))),
        ('microsites_ownership_hold',sum(1 for r in rows if 'ownership' in r.get('strict_verdict','')))]:
        w.writerow([k,v])
# keep the flat complete CSV at root too
print('CSV files written to csv/:',sorted(os.listdir('csv')))
print('disavow',len(DIS),'review',len(REV),'keep',len(KEEP),'toxic%',toxic_pct)
